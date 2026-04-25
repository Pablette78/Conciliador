from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ─── Estilos ───────────────────────────────────────────────────────────────────
BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
DF  = Font(name='Arial', size=10)
TF  = Font(name='Arial', bold=True, size=10)
TFW = Font(name='Arial', bold=True, size=10, color='FFFFFF')
CA  = Alignment(horizontal='center')

AZUL_HEADER = PatternFill('solid', fgColor='2F5496')
AZUL_CLARO  = PatternFill('solid', fgColor='D6E4F0')
VERDE       = PatternFill('solid', fgColor='C6EFCE')
VERDE_TOTAL = PatternFill('solid', fgColor='E2EFDA')
NARANJA     = PatternFill('solid', fgColor='FCE4D6')
AMARILLO    = PatternFill('solid', fgColor='FFEB9C')
ROJO        = PatternFill('solid', fgColor='FFC7CE')
ROJO_FUERTE = PatternFill('solid', fgColor='FF0000')
CELESTE_BOX = PatternFill('solid', fgColor='DDEEFF')
GRIS_CLARO  = PatternFill('solid', fgColor='F2F2F2')

MF  = '#,##0.00'
DTF = 'DD/MM/YYYY'

SH_BANCO = "1. Auditoría Banco"
SH_SIST  = "2. Auditoría Sistema"
SH_IMP   = "3. Detalle Impuestos"
SH_CONC  = "4. Conciliación"

def _h(ws, r, headers, start_col=1):
    for i, h in enumerate(headers):
        col = start_col + i
        cl = ws.cell(row=r, column=col, value=h)
        cl.font = TFW; cl.fill = AZUL_HEADER; cl.alignment = CA; cl.border = BD

def _sf(ws, r, mc, font=None, fill=None):
    for c in range(1, mc + 1):
        cl = ws.cell(row=r, column=c)
        if font: cl.font = font
        if fill: cl.fill = fill
        cl.border = BD

def _crear_auditoria_banco(wb, datos_banco, mes_anio):
    ws = wb.create_sheet(SH_BANCO)
    ws['A1'] = f'AUDITORÍA BANCO - {mes_anio}'; ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = 'Saldo Anterior:'; ws['D3'] = datos_banco.saldo_anterior or 0; ws['D3'].number_format = MF
    _h(ws, 5, ['Fecha', 'Concepto', 'Débito', 'Crédito', 'Detalle', 'Motor'])
    r = 6
    for m in datos_banco.movimientos:
        ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=m.concepto)
        if m.debito: ws.cell(row=r, column=3, value=m.debito).number_format = MF
        if m.credito: ws.cell(row=r, column=4, value=m.credito).number_format = MF
        ws.cell(row=r, column=5, value=m.descripcion); ws.cell(row=r, column=6, value=m.tipo)
        _sf(ws, r, 6, DF); r += 1
    ws.cell(row=r, column=2, value='TOTALES'); ws.cell(row=r, column=3, value=f'=SUM(C6:C{r-1})').number_format = MF; ws.cell(row=r, column=4, value=f'=SUM(D6:D{r-1})').number_format = MF
    row_tot = r; r += 1
    ws.cell(row=r, column=1, value='Final Extracto:'); ws.cell(row=r, column=4, value=datos_banco.saldo_final or 0).number_format = MF
    row_sf = r
    for col, w in [('A', 14), ('B', 30), ('C', 15), ('D', 15), ('E', 30), ('F', 20)]: ws.column_dimensions[col].width = w
    return {'row_saldo_ini': 3, 'row_totales': row_tot, 'row_saldo_fin': row_sf}

def _crear_auditoria_sistema(wb, movs_sist, mes_anio):
    ws = wb.create_sheet(SH_SIST)
    ws['A1'] = f'AUDITORÍA SISTEMA - {mes_anio}'; ws['A1'].font = Font(bold=True, size=14)
    _h(ws, 3, ['Fecha', 'Ref', 'Concepto', 'Detalle', 'Debe', 'Haber'])
    r = 4
    for m in movs_sist:
        ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=m.referencia); ws.cell(row=r, column=3, value=m.concepto)
        if m.debito: ws.cell(row=r, column=5, value=m.debito).number_format = MF
        if m.credito: ws.cell(row=r, column=6, value=m.credito).number_format = MF
        _sf(ws, r, 6, DF); r += 1
    ws.cell(row=r, column=4, value='TOTALES'); ws.cell(row=r, column=5, value=f'=SUM(E4:E{r-1})').number_format = MF; ws.cell(row=r, column=6, value=f'=SUM(F4:F{r-1})').number_format = MF
    row_tot = r
    for col, w in [('A', 14), ('B', 20), ('C', 30), ('D', 30), ('E', 15), ('F', 15)]: ws.column_dimensions[col].width = w
    return {'row_totales': row_tot}

def _crear_detalle_impuestos(wb, resultado, mes_anio):
    ws = wb.create_sheet(SH_IMP)
    ws['A1'] = f'GASTOS E IMPUESTOS - {mes_anio}'; ws['A1'].font = Font(bold=True, size=14)
    r = 3; rows_sub = []
    for cat, datos in sorted(resultado.gastos_por_categoria.items()):
        if not datos['items']: continue
        cl = ws.cell(row=r, column=1, value=cat.replace('_', ' ')); cl.font = TFW; cl.fill = AZUL_HEADER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
        _h(ws, r, ['Fecha', 'Concepto', 'Ref', 'Monto']); r += 1
        start = r
        for m in sorted(datos['items'], key=lambda x: x.fecha):
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF; ws.cell(row=r, column=2, value=m.concepto)
            ws.cell(row=r, column=4, value=m.debito or m.credito).number_format = MF; _sf(ws, r, 4, DF); r += 1
        ws.cell(row=r, column=3, value='SUBTOTAL'); ws.cell(row=r, column=4, value=f'=SUM(D{start}:D{r-1})').number_format = MF; rows_sub.append(r); r += 2
    ws.cell(row=r, column=1, value='TOTAL GENERAL'); ws.cell(row=r, column=4, value='+'.join(f'D{s}' for s in rows_sub) if rows_sub else '0').number_format = MF
    row_tot = r
    for col, w in [('A', 14), ('B', 40), ('C', 30), ('D', 15)]: ws.column_dimensions[col].width = w
    return {'row_gran_total': row_tot}

def _crear_conciliacion(wb, resultado, mes_anio):
    ws = wb.create_sheet(SH_CONC)
    ws['A1'] = f'HOJA DE CONCILIACIÓN - {mes_anio}'; ws['A1'].font = Font(bold=True, size=14)
    _h(ws, 3, ['Fecha Bco', 'Déb Bco', 'Cré Bco', 'Concepto Bco', 'Fecha Sist', 'Debe Sist', 'Haber Sist', 'Ref Sist', 'Dif $', 'Estado', 'Dif Cruce'])
    r = 4; row_ini = r
    for it in resultado.conciliados:
        b, s = it.banco, it.sistema
        ws.cell(row=r, column=1, value=b.fecha).number_format = DTF
        if b.debito: ws.cell(row=r, column=2, value=b.debito).number_format = MF
        if b.credito: ws.cell(row=r, column=3, value=b.credito).number_format = MF
        ws.cell(row=r, column=4, value=b.concepto); ws.cell(row=r, column=5, value=s.fecha).number_format = DTF
        if s.debito: ws.cell(row=r, column=6, value=s.debito).number_format = MF
        if s.credito: ws.cell(row=r, column=7, value=s.credito).number_format = MF
        ws.cell(row=r, column=8, value=s.referencia or s.concepto)
        ws.cell(row=r, column=9, value=getattr(it, 'diferencia', 0)).number_format = MF
        ws.cell(row=r, column=10, value=it.estado); ws.cell(row=r, column=11, value=f'=+B{r}-G{r}-C{r}+F{r}').number_format = MF
        _sf(ws, r, 11, DF, VERDE if it.estado == 'CONCILIADO' else ROJO); r += 1
    row_fin = r - 1
    r += 1; row_sb = None
    if resultado.solo_banco:
        ws.cell(row=r, column=1, value='SOLO EN BANCO'); r += 1
        _h(ws, r, ['Fecha', 'Concepto', 'Tipo', 'Débito', 'Crédito']); r += 1
        start = r
        for m in resultado.solo_banco:
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF; ws.cell(row=r, column=2, value=m.concepto)
            if m.debito: ws.cell(row=r, column=4, value=m.debito).number_format = MF
            if m.credito: ws.cell(row=r, column=5, value=m.credito).number_format = MF
            _sf(ws, r, 5, DF, NARANJA); r += 1
        row_sb = r; r += 1
    r += 1; row_ss = None
    if resultado.solo_sistema:
        ws.cell(row=r, column=1, value='SOLO EN SISTEMA'); r += 1
        _h(ws, r, ['Fecha', 'Ref', 'Concepto', 'Debe', 'Haber']); r += 1
        start = r
        for m in resultado.solo_sistema:
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            if m.debito: ws.cell(row=r, column=4, value=m.debito).number_format = MF
            if m.credito: ws.cell(row=r, column=5, value=m.credito).number_format = MF
            _sf(ws, r, 5, DF, NARANJA); r += 1
        row_ss = r; r += 1
    return {'row_ini': row_ini, 'row_fin': row_fin, 'row_sb': row_sb, 'row_ss': row_ss}

def _llenar_resumen_clean(ws, resultado, datos_banco, mes_anio, rb, rs, ri, rc):
    ws.title = "Resumen"
    ws.merge_cells('A1:G1'); ws['A1'] = f'RESUMEN CONCILIACIÓN - {mes_anio}'; ws['A1'].font = Font(bold=True, size=16); ws['A1'].alignment = CA
    def qf(ws, r, lbl, f, font=DF, fill=None):
        ws.cell(row=r, column=1, value=lbl).font = font; ws.cell(row=r, column=1).border = BD
        if fill: ws.cell(row=r, column=1).fill = fill
        cc = ws.cell(row=r, column=2, value=f); cc.number_format = MF; cc.font = font; cc.border = BD
        if fill: cc.fill = fill
    r = 3; ws.merge_cells('A3:B3'); cl = ws.cell(row=3, column=1, value='CONCILIACIÓN'); cl.fill = AZUL_HEADER; cl.font = TFW; cl.alignment = CA
    r = 4; qf(ws, r, 'Saldo Inicial (Extracto)', f"=+'{SH_BANCO}'!D{rb['row_saldo_ini']}", TF, AZUL_CLARO); r += 1
    qf(ws, r, '(+) Total Debe Mayor', f"=+'{SH_SIST}'!E{rs['row_totales']}" if 'row_totales' in rs else "=0"); r += 1
    qf(ws, r, '(-) Total Haber Mayor', f"=+'{SH_SIST}'!F{rs['row_totales']}" if 'row_totales' in rs else "=0"); r += 1
    row_calc = r; qf(ws, r, '(=) Saldo Final Libro (Calc)', f'=+B4+B5-B6', TF, AZUL_CLARO); r += 1
    r += 1; ws.merge_cells(f'A{r}:B{r}'); ws.cell(row=r, column=1, value='AJUSTES PENDIENTES').font = TF; ws.cell(row=r, column=1).alignment = CA; r += 1
    row_sb_c = r; qf(ws, r, '(+) Banco Créditos Pend', f"=+'{SH_CONC}'!E{rc['row_sb']}" if rc['row_sb'] else "=0"); r += 1
    row_sb_d = r; qf(ws, r, '(-) Banco Débitos Pend', f"=+'{SH_CONC}'!D{rc['row_sb']}" if rc['row_sb'] else "=0"); r += 1
    row_ss_d = r; qf(ws, r, '(-) Libro Debe Pend (DIT)', f"=+'{SH_CONC}'!D{rc['row_ss']}" if rc['row_ss'] else "=0"); r += 1
    row_ss_h = r; qf(ws, r, '(+) Libro Haber Pend (CHQP)', f"=+'{SH_CONC}'!E{rc['row_ss']}" if rc['row_ss'] else "=0"); r += 1
    row_imp = r; qf(ws, r, '(-) Impuestos y Gastos', f"=+'{SH_IMP}'!D{ri['row_gran_total']}"); r += 1
    r += 1; qf(ws, r, '(=) SALDO FINAL AJUSTADO', f'=+B{row_calc}+B{row_sb_c}-B{row_sb_d}-B{row_ss_d}+B{row_ss_h}-B{row_imp}', TFW, AZUL_HEADER); row_ajuste = r; r += 1
    qf(ws, r, 'Saldo Real Extracto', f"=+'{SH_BANCO}'!D{rb['row_saldo_fin']}", TF, GRIS_CLARO); row_real = r; r += 1
    ws.cell(row=r, column=1, value='DIFERENCIA (Debe ser 0)').font = TF; ws.cell(row=r, column=1).border = BD
    cfinal = ws.cell(row=r, column=2, value=f'=+B{row_real}-B{row_ajuste}'); cfinal.number_format = MF; cfinal.font = TF; cfinal.border = BD; cfinal.fill = VERDE_TOTAL
    
    col_d, col_e = 4, 5
    r_v = 3; ws.merge_cells(start_row=r_v, start_column=col_d, end_row=r_v, end_column=7); cl = ws.cell(row=r_v, column=col_d, value='VERIFICACIÓN PDF'); cl.fill = AZUL_HEADER; cl.font = TFW; cl.alignment = CA
    for i, (l, f) in enumerate([('Inicio', f"=+'{SH_BANCO}'!D{rb['row_saldo_ini']}"), ('Créditos (+)', f"=+'{SH_BANCO}'!D{rb['row_totales']}"), ('Débitos (-)', f"=+'{SH_BANCO}'!C{rb['row_totales']}"), ('Calculado', '=+E5+E6-E7'), ('Real PDF', f"=+'{SH_BANCO}'!D{rb['row_saldo_fin']}"), ('Dif.', '=+E8-E9')]):
        row_v = 4 + i; ws.cell(row=row_v, column=col_d, value=l).border = BD
        ws.merge_cells(start_row=row_v, start_column=col_e, end_row=row_v, end_column=6)
        cc = ws.cell(row=row_v, column=col_e, value=f); cc.number_format = MF; cc.border = BD
        if i == 5: cc.fill = VERDE_TOTAL

    r_audit = 12; ws.merge_cells(start_row=r_audit, start_column=col_d, end_row=r_audit, end_column=7); cl = ws.cell(row=r_audit, column=col_d, value='AUDITORÍA CRUZADA'); cl.fill = AZUL_HEADER; cl.font = TFW; cl.alignment = CA
    r_audit += 1; _h(ws, r_audit, ['Control', 'Suma Items', 'Valor Ref', 'Dif'], start_col=col_d)
    for l, fi, fr in [('Bco Déb', f"=+B{row_sb_d}+B{row_imp}+SUM('{SH_CONC}'!B{rc['row_ini']}:B{rc['row_fin']})", f"='{SH_BANCO}'!C{rb['row_totales']}"), ('Bco Cré', f"=+B{row_sb_c}+SUM('{SH_CONC}'!C{rc['row_ini']}:C{rc['row_fin']})", f"='{SH_BANCO}'!D{rb['row_totales']}"), ('Sist Debe', f"=+B{row_ss_d}+SUM('{SH_CONC}'!F{rc['row_ini']}:F{rc['row_fin']})", f"='{SH_SIST}'!E{rs['row_totales']}" if 'row_totales' in rs else '0'), ('Sist Hab', f"=+B{row_ss_h}+SUM('{SH_CONC}'!G{rc['row_ini']}:G{rc['row_fin']})", f"='{SH_SIST}'!F{rs['row_totales']}" if 'row_totales' in rs else '0')]:
        r_audit += 1; ws.cell(row=r_audit, column=col_d, value=lbl).border = BD
        ws.cell(row=r_audit, column=5, value=fi).number_format = MF; ws.cell(row=r_audit, column=5).border = BD
        ws.cell(row=r_audit, column=6, value=fr).number_format = MF; ws.cell(row=r_audit, column=6).border = BD
        cd = ws.cell(row=r_audit, column=7, value=f'=+E{r_audit}-F{r_audit}'); cd.number_format = MF; cd.border = BD; cd.fill = VERDE_TOTAL

    r_g = 18; ws.merge_cells(f'A{r_g}:B{r_g}'); cl = ws.cell(row=r_g, column=1, value='GASTOS RESUMEN'); cl.fill = AZUL_HEADER; cl.font = TFW; cl.alignment = CA; r_g += 1
    for c, d in sorted(resultado.gastos_por_categoria.items()):
        ws.cell(row=r_g, column=1, value=c.replace('_',' ')).border = BD; ws.cell(row=r_g, column=2, value=d['total']).number_format = MF; ws.cell(row=r_g, column=2).border = BD; r_g += 1
    ws.column_dimensions['A'].width = 38; ws.column_dimensions['B'].width = 18; ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 18; ws.column_dimensions['F'].width = 18; ws.column_dimensions['G'].width = 12

def generar_excel(resultado, datos_banco, ruta_salida, mes_anio="", movs_sist=None):
    wb = Workbook(); ws_res = wb.active
    rb = _crear_auditoria_banco(wb, datos_banco, mes_anio)
    rs = _crear_auditoria_sistema(wb, movs_sist, mes_anio) if movs_sist else {}
    ri = _crear_detalle_impuestos(wb, resultado, mes_anio)
    rc = _crear_conciliacion(wb, resultado, mes_anio)
    _llenar_resumen_clean(ws_res, resultado, datos_banco, mes_anio, rb, rs, ri, rc)
    wb.save(ruta_salida); return ruta_salida
