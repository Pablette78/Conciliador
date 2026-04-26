"""
GenericExcelBankParser
Parser genérico para extractos bancarios en formato Excel (.xlsx / .xls).
Detecta automáticamente las columnas: Fecha, Concepto, Débito, Crédito (o Importe único).
Sirve como fallback para cualquier banco que exporte su extracto en Excel
y no tenga un parser PDF dedicado.
"""
import os
from datetime import datetime
from .base import BaseParser
from ..models import Movimiento, DatosExtracto


def _leer_hojas_xlsx(ruta):
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True, read_only=True)
    hojas = {}
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        filas = [list(row) for row in ws.iter_rows(values_only=True)]
        hojas[nombre] = filas
    return hojas


def _leer_hojas_xls(ruta):
    import xlrd
    wb = xlrd.open_workbook(ruta)
    hojas = {}
    for nombre in wb.sheet_names():
        ws = wb.sheet_by_name(nombre)
        filas = []
        for r in range(ws.nrows):
            fila = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                valor = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        t = xlrd.xldate_as_tuple(valor, wb.datemode)
                        valor = datetime(t[0], t[1], t[2])
                    except Exception:
                        pass
                fila.append(valor)
            filas.append(fila)
        hojas[nombre] = filas
    return hojas


def _texto(val):
    """Convierte un valor de celda a string en mayúsculas para comparación."""
    return str(val).upper().strip() if val is not None else ''


def _mapear_columnas_banco(fila):
    """
    Detecta las columnas de fecha, concepto, débito, crédito e importe único.
    Retorna un dict con los índices de columna detectados.
    """
    col = {}
    NOMBRES_FECHA = ('FECHA',)
    NOMBRES_CONCEPTO = ('CONCEPTO', 'DESCRIPCION', 'DESCRIPCIÓN', 'DETALLE', 'MOVIMIENTO', 'TRANSACCION', 'TRANSACCIÓN')
    NOMBRES_DEBITO = ('DEBITO', 'DÉBITO', 'DÉBITOS', 'DEBITOS', 'CARGO', 'CARGOS', 'SALIDA', 'SALIDAS', 'EGR', 'EGR.')
    NOMBRES_CREDITO = ('CREDITO', 'CRÉDITO', 'CRÉDITOS', 'CREDITOS', 'ABONO', 'ABONOS', 'ENTRADA', 'ENTRADAS', 'ACR', 'ACRED')
    NOMBRES_IMPORTE = ('IMPORTE', 'MONTO', 'AMOUNT')
    NOMBRES_SALDO = ('SALDO', 'BALANCE')

    for i, val in enumerate(fila):
        t = _texto(val)
        if not t:
            continue
        if any(n in t for n in NOMBRES_FECHA) and 'fecha' not in col:
            col['fecha'] = i
        if any(t == n for n in NOMBRES_CONCEPTO) and 'concepto' not in col:
            col['concepto'] = i
        if any(n in t for n in NOMBRES_CONCEPTO) and 'concepto' not in col:
            col['concepto'] = i
        if any(t == n for n in NOMBRES_DEBITO) and 'debito' not in col:
            col['debito'] = i
        if any(t == n for n in NOMBRES_CREDITO) and 'credito' not in col:
            col['credito'] = i
        if any(t == n for n in NOMBRES_IMPORTE) and 'importe' not in col:
            col['importe'] = i
        if any(t == n for n in NOMBRES_SALDO) and 'saldo' not in col:
            col['saldo'] = i

    return col


def _parse_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace('.', '').replace(',', '.').replace('$', '').replace(' ', '')
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _parse_fecha(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    if isinstance(val, float):
        try:
            import xlrd
            t = xlrd.xldate_as_tuple(val, 0)
            return datetime(t[0], t[1], t[2])
        except Exception:
            pass
    return None


class GenericExcelBankParser(BaseParser):
    """
    Parser genérico para extractos bancarios en Excel.
    Detecta automáticamente los encabezados y mapea las columnas.
    Soporta formato con columnas separadas (Débito / Crédito) o
    formato con columna única de Importe (con signo o sin signo).
    """

    NOMBRE_BANCO = "Banco (Excel)"

    def parse(self, ruta_archivo: str) -> DatosExtracto:
        ext = os.path.splitext(ruta_archivo)[1].lower()
        if ext == '.xls':
            hojas = _leer_hojas_xls(ruta_archivo)
        elif ext in ('.xlsx', '.xlsm'):
            hojas = _leer_hojas_xlsx(ruta_archivo)
        else:
            raise Exception(f"Formato no soportado: {ext}")

        # Elegir la hoja con más filas (la de datos)
        nombre_hoja = max(hojas, key=lambda n: len(hojas[n]))
        filas = hojas[nombre_hoja]

        # Buscar fila de encabezados en las primeras 10 filas
        col_map = {}
        header_idx = None
        for i, fila in enumerate(filas[:10]):
            cm = _mapear_columnas_banco(fila)
            if 'fecha' in cm and ('debito' in cm or 'credito' in cm or 'importe' in cm):
                col_map = cm
                header_idx = i
                break

        if header_idx is None:
            raise Exception(
                f"No se encontraron encabezados válidos en la hoja '{nombre_hoja}'.\n"
                f"El archivo debe tener columnas: Fecha, Débito/Crédito (o Importe)."
            )

        movimientos = []
        saldo_running = None

        for fila in filas[header_idx + 1:]:
            if not any(v is not None for v in fila):
                continue

            fecha_val = fila[col_map['fecha']] if col_map.get('fecha') is not None and col_map['fecha'] < len(fila) else None
            fecha = _parse_fecha(fecha_val)
            if not fecha:
                continue

            concepto_idx = col_map.get('concepto')
            concepto = str(fila[concepto_idx]).strip() if concepto_idx is not None and concepto_idx < len(fila) and fila[concepto_idx] else ''

            debito = 0.0
            credito = 0.0

            if 'debito' in col_map and 'credito' in col_map:
                # Formato con dos columnas separadas
                deb_idx = col_map['debito']
                cred_idx = col_map['credito']
                debito  = _parse_float(fila[deb_idx]  if deb_idx  < len(fila) else None)
                credito = _parse_float(fila[cred_idx] if cred_idx < len(fila) else None)

            elif 'importe' in col_map:
                # Formato con columna única de importe.
                # Si tiene columna de saldo, deducimos dirección matemáticamente.
                imp_idx = col_map['importe']
                importe = _parse_float(fila[imp_idx] if imp_idx < len(fila) else None)
                monto = abs(importe)

                saldo_idx = col_map.get('saldo')
                if saldo_idx is not None and saldo_idx < len(fila):
                    saldo_nuevo = _parse_float(fila[saldo_idx])
                    if saldo_running is not None:
                        if saldo_nuevo > saldo_running:
                            credito = monto
                        else:
                            debito = monto
                    else:
                        # Sin historial de saldo: usar signo del importe
                        if importe < 0:
                            debito = monto
                        else:
                            credito = monto
                    saldo_running = saldo_nuevo
                else:
                    # Sin columna de saldo: usar signo del importe
                    if importe < 0:
                        debito = monto
                    else:
                        credito = monto

            if debito == 0 and credito == 0:
                continue

            tipo = self.clasificar_concepto(concepto)

            movimientos.append(Movimiento(
                fecha=fecha,
                concepto=concepto,
                debito=debito,
                credito=credito,
                tipo=tipo,
                descripcion=""
            ))

        return DatosExtracto(
            banco=self.NOMBRE_BANCO,
            titular="",
            movimientos=movimientos,
            saldo_anterior=0.0,
            saldo_final=0.0
        )
