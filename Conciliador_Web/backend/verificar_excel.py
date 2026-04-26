import os
import sys
from datetime import datetime
from openpyxl import load_workbook

# Mocking the models and result for the test
class MockMovimiento:
    def __init__(self, fecha, concepto, debito=0, credito=0, tipo='VARIOS', descripcion=''):
        self.fecha = fecha
        self.concepto = concepto
        self.debito = debito
        self.credito = credito
        self.tipo = tipo
        self.descripcion = descripcion

class MockResultado:
    def __init__(self):
        self.conciliados = []
        self.solo_banco = [MockMovimiento(datetime(2026,2,10), 'EXTRACCION CAJERO', 5000, 0, 'GASTO')]
        self.solo_sistema = [MockMovimiento(datetime(2026,2,15), 'CHEQUE EMITIDO', 0, 12000, 'SISTEMA', 'CH 123')]
        self.gastos_por_categoria = {
            'COMISION': {'total': 500, 'items': [MockMovimiento(datetime(2026,2,1), 'COMISION MANT', 500, 0, 'GASTO')]},
            'IVA': {'total': 105, 'items': [MockMovimiento(datetime(2026,2,1), 'IVA 21%', 105, 0, 'GASTO')]}
        }
        self.validación_saldos = {'saldo_final_calculado': 1000, 'saldo_final_extracto': 1000}

class MockDatosBanco:
    def __init__(self):
        self.titular = "EMPRESA TEST S.A."
        self.saldo_anterior = 50000
        self.saldo_final = 50000 + 0 - 605 - 5000 # Simplificado
        self.movimientos = [
             MockMovimiento(datetime(2026,2,1), 'COMISION MANT', 500, 0, 'GASTO'),
             MockMovimiento(datetime(2026,2,1), 'IVA 21%', 105, 0, 'GASTO'),
             MockMovimiento(datetime(2026,2,10), 'EXTRACCION CAJERO', 5000, 0, 'GASTO')
        ]

# Import the actual generator
sys.path.append(os.getcwd())
try:
    from generador_excel import generar_excel
    
    res = MockResultado()
    db = MockDatosBanco()
    output = "test_resumen_clean.xlsx"
    
    generar_excel(res, db, output, "FEB-2026")
    
    # Now read it back to show the structure
    wb = load_workbook(output, data_only=True)
    ws = wb["Resumen"]
    
    print("| Fila | Col A (Contenido) | Col B (Monto) | Col D (Auditoría) | Col E (Valor) |")
    print("|---|---|---|---|---|")
    for r in range(1, 26):
        a = ws.cell(row=r, column=1).value or ""
        b = ws.cell(row=r, column=2).value or ""
        d = ws.cell(row=r, column=4).value or ""
        e = ws.cell(row=r, column=5).value or ""
        
        # Format values to match Excel look
        if isinstance(b, (int, float)): b = f"{b:,.2f}"
        if isinstance(e, (int, float)): e = f"{e:,.2f}"
        
        print(f"| {r} | {a} | {b} | {d} | {e} |")

except Exception as ex:
    print(f"Error: {ex}")
