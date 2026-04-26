"""
Parser para el archivo Excel del sistema contable.
Lee los movimientos del Mayor y los normaliza.
Soporta formatos: .xls (Excel 97-2003) y .xlsx (Excel moderno)
"""
import os
from datetime import datetime


def _leer_hojas_xls(ruta):
    """Lee un archivo .xls con xlrd y devuelve dict de hojas con sus filas"""
    import xlrd
    wb = xlrd.open_workbook(ruta)
    hojas = {}
    for nombre in wb.sheet_names():
        ws = wb.sheet_by_name(nombre)
        filas = []
        for row_idx in range(ws.nrows):
            fila = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                valor = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        tupla = xlrd.xldate_as_tuple(valor, wb.datemode)
                        valor = datetime(tupla[0], tupla[1], tupla[2])
                    except Exception:
                        pass
                fila.append(valor)
            filas.append(fila)
        hojas[nombre] = filas
    return hojas


def _leer_hojas_xlsx(ruta):
    """Lee un archivo .xlsx con openpyxl y devuelve dict de hojas con sus filas"""
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True)
    hojas = {}
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        filas = []
        for row in ws.iter_rows(values_only=True):
            filas.append(list(row))
        hojas[nombre] = filas
    return hojas


def _buscar_hoja_mayor(hojas):
    """Busca la hoja que contenga los movimientos del mayor"""
    for nombre, filas in hojas.items():
        for fila in filas[:20]: # Escanear más profundo para encontrar la hoja correcta
            textos = ' '.join(str(c).upper() if c else '' for c in fila)
            if 'FECHA' in textos and \
               any(x in textos for x in ['DEBE', 'HABER', 'TOTAL', 'IMPORTE', 'VALOR', 'MONTO', 'DOCUMENTO']):
                return nombre
    return max(hojas, key=lambda n: len(hojas[n]))


def _mapear_columnas(fila):
    """Identifica qué columna corresponde a cada campo"""
    col_map = {}
    for col_idx, valor in enumerate(fila):
        texto = str(valor).upper().strip() if valor else ''
        if 'FECHA' in texto and 'fecha' not in col_map:
            col_map['fecha'] = col_idx
        if ('DOCUMENTO' in texto or texto == 'DOC') and 'documento' not in col_map:
            col_map['documento'] = col_idx
        if any(x in texto for x in ['DESCRIPCI', 'CUENTA', 'CONCEPTO', 'DETALLE']):
            if 'descripcion' not in col_map:
                col_map['descripcion'] = col_idx
        if texto == 'DETALLE' and 'detalle' not in col_map:
            col_map['detalle'] = col_idx
        if any(x in texto for x in ['OPERACI']):
            if 'operacion' not in col_map:
                col_map['operacion'] = col_idx
        if 'RAZON SOCIAL' in texto or 'ACLARACION' in texto:
            if 'razon_social' not in col_map:
                col_map['razon_social'] = col_idx
        # Columna de plata que ENTRA (Debe, Ing., Ingreso, Crédito, etc.)
        NOMBRES_ENTRADA = (
            'DEBE', 'ING.', 'ING', 'INGRESO', 'INGRESOS',
            'CREDITO', 'CRÉDITO', 'CREDITOS', 'CRÉDITOS',
        )
        if texto in NOMBRES_ENTRADA and 'debe' not in col_map:
            col_map['debe'] = col_idx
        # Columna de plata que SALE (Haber, Egr., Egreso, Débito, etc.)
        NOMBRES_SALIDA = (
            'HABER', 'EGR.', 'EGR', 'EGRESO', 'EGRESOS',
            'DEBITO', 'DÉBITO', 'DEBITOS', 'DÉBITOS',
        )
        if texto in NOMBRES_SALIDA and 'haber' not in col_map:
            col_map['haber'] = col_idx
        # Columna de plata genérica (Total o Importe) si faltan Debe/Haber
        if any(x in texto for x in ['TOTAL', 'IMPORTE', 'VALOR', 'MONTO']) and 'haber' not in col_map:
            col_map['haber'] = col_idx
        if 'SALDO' in texto and 'saldo' not in col_map:
            col_map['saldo'] = col_idx
    return col_map


def _obtener_valor(fila, col_map, campo, default=''):
    """Obtiene un valor de la fila según el mapeo de columnas"""
    idx = col_map.get(campo)
    if idx is None or idx >= len(fila):
        return default
    valor = fila[idx]
    return valor if valor is not None else default


from core.models import Movimiento

def parsear_excel(ruta_excel):
    """
    Lee el Excel del sistema contable y extrae los movimientos.
    Soporta .xls y .xlsx automáticamente.
    Busca la hoja con movimientos (Mayor) automáticamente.

    Retorna lista de objetos Movimiento
    """
    extension = os.path.splitext(ruta_excel)[1].lower()

    if extension == '.xls':
        hojas = _leer_hojas_xls(ruta_excel)
    elif extension in ('.xlsx', '.xlsm'):
        hojas = _leer_hojas_xlsx(ruta_excel)
    else:
        raise Exception(f"Formato no soportado: {extension}. Usá .xls o .xlsx")

    nombre_hoja = _buscar_hoja_mayor(hojas)
    filas = hojas[nombre_hoja]

    # Encontrar fila de encabezados
    header_row_idx = None
    col_map = {}
    for idx, fila in enumerate(filas[:20]): # Aumentado a 20 para archivos con muchos encabezados
        col_map = _mapear_columnas(fila)
        if 'fecha' in col_map and ('debe' in col_map or 'haber' in col_map):
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise Exception(
            f"No se encontraron encabezados válidos en la hoja '{nombre_hoja}'.\n"
            f"El archivo debe tener columnas: Fecha, Debe/Haber o Total"
        )

    movimientos = []
    for fila in filas[header_row_idx + 1:]:
        fecha = _obtener_valor(fila, col_map, 'fecha')
        if not fecha:
            continue

        # Convertir fecha string a datetime
        if isinstance(fecha, str):
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    fecha = datetime.strptime(fecha.strip(), fmt)
                    break
                except ValueError:
                    continue

        # xlrd puede devolver float para fechas mal leídas
        if isinstance(fecha, float):
            try:
                import xlrd
                tupla = xlrd.xldate_as_tuple(fecha, 0)
                fecha = datetime(tupla[0], tupla[1], tupla[2])
            except Exception:
                continue

        if not isinstance(fecha, datetime):
            continue

        debe = _obtener_valor(fila, col_map, 'debe', 0) or 0
        haber = _obtener_valor(fila, col_map, 'haber', 0) or 0

        try:
            debe = float(debe)
            haber = float(haber)
        except (ValueError, TypeError):
            continue

        doc = str(_obtener_valor(fila, col_map, 'documento', ''))
        desc = str(_obtener_valor(fila, col_map, 'descripcion', ''))
        detalle = str(_obtener_valor(fila, col_map, 'detalle', ''))
        razon = str(_obtener_valor(fila, col_map, 'razon_social', ''))
        operacion = str(_obtener_valor(fila, col_map, 'operacion', ''))

        # Combinar detalle con razón social si existe
        if razon and razon != detalle:
            detalle = f"{detalle} - {razon}".strip(' -')

        if 'SALDO INICIAL' in desc.upper() or 'SALDO INICIAL' in doc.upper():
            continue

        if debe == 0 and haber == 0:
            continue

        movimientos.append(Movimiento(
            fecha=fecha,
            referencia=doc.strip() or operacion.strip(),
            concepto=desc.strip(),
            descripcion=detalle.strip(),
            debito=debe,
            credito=haber,
            tipo='SISTEMA'
        ))

    return movimientos
