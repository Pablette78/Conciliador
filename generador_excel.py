"""
Generador del Excel de conciliación.
Hojas:
  Resumen
  1. Auditoría Banco
  2. Auditoría Sistema 
  3. Detalle Impuestos
  4. Conciliación
  5. Conceptos Agrupados
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Estilos
HF = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HB = PatternFill('solid', fgColor='2F5496')
SF = Font(name='Arial', bold=True, size=10)
SB = PatternFill('solid', fgColor='D6E4F0')
DF = Font(name='Arial', size=10)
TF = Font(name='Arial', bold=True, size=10)
TB = PatternFill('solid', fgColor='E2EFDA')
MB = PatternFill('solid', fgColor='E2EFDA')
NB = PatternFill('solid', fgColor='FCE4D6')
BB = PatternFill('solid', fgColor='BDD7EE')
MF = '#,##0.00'
DTF = 'DD/MM/YYYY'
BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CA = Alignment(horizontal='center')

def _h(ws, r, headers):
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row=r, column=c, value=h)
        cl.font = HF; cl.fill = HB; cl.alignment = CA; cl.border = BD

def _sf(ws, r, mc, font=None, fill=None):
    for c in range(1, mc + 1):
        cl = ws.cell(row=r, column=c)
        if font: cl.font = font
        if fill: cl.fill = fill
        cl.border = BD

def _titulo(ws, r, texto, mc=7):
    ws.cell(row=r, column=1, value=texto).font = Font(name='Arial', bold=True, size=12, color='2F5496')

# ============================================================
# HOJA 0: RESUMEN
# ============================================================
def _crear_resumen(wb, resultado, datos_banco, mes_anio):
    ws = wb.active
    ws.title = "Resumen"

    ws.merge_cells('A1:D1')
    ws['A1'] = f'RESUMEN CONCILIACIÓN - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color='2F5496')
    ws['A2'] = datos_banco.titular
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    r = 4
    v = resultado.validación_saldos
    items = [
        ('Saldo Anterior s/ Banco', v.get('saldo_anterior', 0)),
        ('(+) Total Créditos', v.get('total_creditos', 0)),
        ('(-) Total Débitos', v.get('total_debitos', 0)),
        ('(=) Saldo Final Calculado', v.get('saldo_final_calculado', 0)),
        ('Saldo Final s/ Banco (Extracto)', v.get('saldo_final_extracto', 0)),
    ]
    for desc, val in items:
        ws.cell(row=r, column=1, value=desc).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=val).number_format = MF
        ws.cell(row=r, column=2).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=1).fill = BB
        ws.cell(row=r, column=2).fill = BB
        r += 1

    if not v.get('coincide', True):
        ws.cell(row=r, column=1, value="⚠ DIFERENCIA EN SALDO FINAL").font = Font(name='Arial', bold=True, color='FF0000')
        ws.cell(row=r, column=2, value=v.get('saldo_final_extracto', 0) - v.get('saldo_final_calculado', 0)).number_format = MF
    r += 2

    ws.cell(row=r, column=1, value='CONCILIACIÓN').font = Font(name='Arial', bold=True, size=12, color='2F5496')
    r += 1

    total_conc = len(resultado.conciliados)
    con_diff = sum(1 for c in resultado.conciliados if c.diferencia != 0)
    
    stats = [
        ('Movimientos conciliados', total_conc),
        ('Con diferencia de monto', con_diff),
        ('Solo en banco (pendientes)', len(resultado.solo_banco)),
        ('Solo en sistema (pendientes)', len(resultado.solo_sistema)),
    ]

    for desc, val in stats:
        ws.cell(row=r, column=1, value=desc).font = DF
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).font = DF
        ws.cell(row=r, column=2).border = BD
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='GASTOS BANCARIOS E IMPUESTOS').font = Font(name='Arial', bold=True, size=12, color='2F5496')
    r += 1

    gran_total = 0
    for cat, datos in resultado.gastos_por_categoria.items():
        nombre = cat
        ws.cell(row=r, column=1, value=nombre).font = DF
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=datos['total']).number_format = MF
        ws.cell(row=r, column=2).font = DF
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=3, value=f"{len(datos['items'])} movimientos").font = Font(name='Arial', size=9, color='888888')
        gran_total += datos['total']
        r += 1

    ws.cell(row=r, column=1, value='TOTAL GASTOS').font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=r, column=1).fill = TB
    ws.cell(row=r, column=1).border = BD
    ws.cell(row=r, column=2, value=gran_total).number_format = MF
    ws.cell(row=r, column=2).font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=r, column=2).fill = TB
    ws.cell(row=r, column=2).border = BD

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20


# ============================================================
# HOJA 1: AUDITORÍA BANCO
# ============================================================
def _crear_auditoria_banco(wb, datos_banco, mes_anio):
    ws = wb.create_sheet("1. Auditoría Banco")

    ws.merge_cells('A1:F1')
    ws['A1'] = f'AUDITORÍA: MOVIMIENTOS DETECTADOS EN BANCO - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A3'] = f'Saldo Anterior detectado: $ {datos_banco.saldo_anterior:,.2f}'
    ws['A3'].font = Font(name='Arial', bold=True, size=10)

    _h(ws, 5, ['Fecha', 'Concepto Original', 'Débito', 'Crédito', 'Descripción/Ref', 'Clasificación Motor'])
    r = 6
    for m in datos_banco.movimientos:
        ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=m.concepto)
        if m.debito: ws.cell(row=r, column=3, value=m.debito).number_format = MF
        if m.credito: ws.cell(row=r, column=4, value=m.credito).number_format = MF
        ws.cell(row=r, column=5, value=m.descripcion)
        ws.cell(row=r, column=6, value=m.tipo)
        if m.tipo != 'OTRO':
            ws.cell(row=r, column=6).font = Font(name='Arial', bold=True, color='9C5700')
        _sf(ws, r, 6, DF)
        r += 1

    ws.cell(row=r, column=2, value='TOTALES').font = TF
    ws.cell(row=r, column=2).fill = TB
    ws.cell(row=r, column=3, value=f'=SUM(C6:C{r-1})').number_format = MF
    ws.cell(row=r, column=3).font = TF; ws.cell(row=r, column=3).fill = TB
    ws.cell(row=r, column=4, value=f'=SUM(D6:D{r-1})').number_format = MF
    ws.cell(row=r, column=4).font = TF; ws.cell(row=r, column=4).fill = TB
    _sf(ws, r, 6)

    for c, w in [('A', 14), ('B', 32), ('C', 16), ('D', 16), ('E', 30), ('F', 24)]:
        ws.column_dimensions[c].width = w

# ============================================================
# HOJA 2: AUDITORÍA SISTEMA
# ============================================================
def _crear_auditoria_sistema(wb, movimientos_sistema, mes_anio):
    ws = wb.create_sheet("2. Auditoría Sistema")

    ws.merge_cells('A1:F1')
    ws['A1'] = f'AUDITORÍA: MOVIMIENTOS DETECTADOS EN SISTEMA - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    _h(ws, 3, ['Fecha', 'Documento/Ref', 'Concepto Original', 'Detalle', 'Debe', 'Haber'])
    r = 4
    for m in movimientos_sistema:
        ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=m.referencia) 
        ws.cell(row=r, column=3, value=m.concepto)    
        ws.cell(row=r, column=4, value=m.descripcion) 
        if m.debito: ws.cell(row=r, column=5, value=m.debito).number_format = MF
        if m.credito: ws.cell(row=r, column=6, value=m.credito).number_format = MF
        _sf(ws, r, 6, DF)
        r += 1

    ws.cell(row=r, column=4, value='TOTALES').font = TF
    ws.cell(row=r, column=4).fill = TB
    ws.cell(row=r, column=5, value=f'=SUM(E4:E{r-1})').number_format = MF
    ws.cell(row=r, column=5).font = TF; ws.cell(row=r, column=5).fill = TB
    ws.cell(row=r, column=6, value=f'=SUM(F4:F{r-1})').number_format = MF
    ws.cell(row=r, column=6).font = TF; ws.cell(row=r, column=6).fill = TB

    for c, w in [('A', 14), ('B', 28), ('C', 32), ('D', 32), ('E', 16), ('F', 16)]:
        ws.column_dimensions[c].width = w

# ============================================================
# HOJA 3: DETALLE DE IMPUESTOS
# ============================================================
def _crear_detalle_impuestos(wb, resultado, mes_anio):
    ws = wb.create_sheet("3. Detalle Impuestos")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f'DETALLE DIARIO DE IMPUESTOS Y GASTOS BANCARIOS - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    
    r = 3
    for cat, datos in sorted(resultado.gastos_por_categoria.items(), key=lambda x: str(x[0])):
        if len(datos['items']) == 0: continue
        
        ws.cell(row=r, column=1, value=str(cat).replace('_', ' ')).font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='4472C4')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        
        _h(ws, r, ['Fecha', 'Concepto Banco', 'Referencia / Detalle', 'Monto'])
        r += 1
        
        start_r = r
        for m in sorted(datos['items'], key=lambda x: x.fecha):
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            ws.cell(row=r, column=2, value=m.concepto[:80]) 
            ws.cell(row=r, column=3, value=m.descripcion[:80])
            ws.cell(row=r, column=4, value=m.debito or m.credito).number_format = MF
            _sf(ws, r, 4, DF)
            r += 1
            
        ws.cell(row=r, column=3, value='TOTAL ' + str(cat)).font = TF
        ws.cell(row=r, column=3).fill = TB
        ws.cell(row=r, column=4, value=f'=SUM(D{start_r}:D{r-1})').number_format = MF
        ws.cell(row=r, column=4).font = TF; ws.cell(row=r, column=4).fill = TB
        _sf(ws, r, 4)
        
        # Agrupar y colapsar las filas de detalle (Outline)
        ws.row_dimensions.group(start_r, r-1, hidden=True)
        r += 2

    for c, w in [('A', 14), ('B', 50), ('C', 45), ('D', 20)]:
        ws.column_dimensions[c].width = w

# ============================================================
# HOJA 4: CONCILIACIÓN
# ============================================================
def _crear_conciliacion(wb, resultado, datos_banco, mes_anio):
    ws = wb.create_sheet("4. Conciliación")

    ws.merge_cells('A1:J1')
    ws['A1'] = f'CONCILIACIÓN BANCARIA - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = datos_banco.titular
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    r = 4
    _titulo(ws, r, '1. CRUCE DE MOVIMIENTOS (Banco vs Sistema)', mc=10)
    r += 1
    _h(ws, r, ['Fecha Banco', 'Monto Banco', 'Conc. Banco', 'Fecha Sist.', 'Monto Sist.', 'Ref. Sist.', 'Dif. $', 'Dif. Días', 'Nivel Cierre', 'Estado'])
    r += 1

    for item in resultado.conciliados:
        b = item.banco
        s = item.sistema
        ws.cell(row=r, column=1, value=b.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=b.monto).number_format = MF
        ws.cell(row=r, column=3, value=b.concepto)
        ws.cell(row=r, column=4, value=s.fecha).number_format = DTF
        ws.cell(row=r, column=5, value=s.monto).number_format = MF
        ws.cell(row=r, column=6, value=s.referencia or s.concepto)
        ws.cell(row=r, column=7, value=getattr(item, 'diferencia', 0)).number_format = MF
        ws.cell(row=r, column=8, value=getattr(item, 'diferencia_dias', 0))
        ws.cell(row=r, column=9, value=item.nivel)
        ws.cell(row=r, column=10, value=item.estado)
        fill = MB if item.estado == 'CONCILIADO' else NB
        _sf(ws, r, 10, DF, fill)
        r += 1

    if resultado.solo_banco:
        r += 1
        ws.cell(row=r, column=1, value=f'SOLO EN BANCO ({len(resultado.solo_banco)} movimientos sin match)').font = Font(name='Arial', bold=True, size=11, color='C00000')
        r += 1; _h(ws, r, ['Fecha', 'Concepto', 'Tipo Automático', 'Crédito', 'Débito', '', '', '']); r += 1
        start_sb = r
        for m in resultado.solo_banco:
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            ws.cell(row=r, column=2, value=m.concepto)
            ws.cell(row=r, column=3, value=m.tipo)
            if m.credito: ws.cell(row=r, column=4, value=m.credito).number_format = MF
            if m.debito: ws.cell(row=r, column=5, value=m.debito).number_format = MF
            _sf(ws, r, 5, DF, NB)
            r += 1
        ws.cell(row=r, column=1, value='TOTAL SOLO EN BANCO').font = TF; ws.cell(row=r, column=1).fill = NB
        ws.cell(row=r, column=4, value=f'=SUM(D{start_sb}:D{r-1})').number_format = MF; ws.cell(row=r, column=4).font = TF; ws.cell(row=r, column=4).fill = NB
        ws.cell(row=r, column=5, value=f'=SUM(E{start_sb}:E{r-1})').number_format = MF; ws.cell(row=r, column=5).font = TF; ws.cell(row=r, column=5).fill = NB
        r += 1

    if resultado.solo_sistema:
        r += 1
        ws.cell(row=r, column=1, value=f'SOLO EN SISTEMA ({len(resultado.solo_sistema)} movimientos sin match)').font = Font(name='Arial', bold=True, size=11, color='C00000')
        r += 1; _h(ws, r, ['Fecha', 'Ref.', 'Concepto', 'Debe', 'Haber', '', '', '']); r += 1
        start_ss = r
        for m in resultado.solo_sistema:
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            ws.cell(row=r, column=2, value=m.referencia)
            ws.cell(row=r, column=3, value=m.concepto)
            if m.debito: ws.cell(row=r, column=4, value=m.debito).number_format = MF
            if m.credito: ws.cell(row=r, column=5, value=m.credito).number_format = MF
            _sf(ws, r, 5, DF, NB)
            r += 1

    for c, w in [('A', 14), ('B', 16), ('C', 30), ('D', 14), ('E', 16), ('F', 25), ('G', 12), ('H', 10), ('I', 15), ('J', 20)]:
        ws.column_dimensions[c].width = w

# ============================================================
# HOJA 5: CONCEPTOS AGRUPADOS
# ============================================================
def _crear_conceptos_agrupados(wb, datos_banco, mes_anio):
    ws = wb.create_sheet("5. Conceptos Agrupados")
    
    ws.merge_cells('A1:E1')
    ws['A1'] = f'AGRUPACIÓN TOTAL POR CONCEPTOS BANCARIOS - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    
    agrup = defaultdict(lambda: {'creditos': 0.0, 'debitos': 0.0, 'cantidad': 0})
    
    for m in datos_banco.movimientos:
        llave = (m.tipo, m.concepto.strip().upper())
        agrup[llave]['creditos'] += m.credito
        agrup[llave]['debitos'] += m.debito
        agrup[llave]['cantidad'] += 1
        
    r = 3
    _h(ws, r, ['Tipo Clasificación', 'Concepto Exacto', 'Cant.', 'Total Débitos', 'Total Créditos'])
    r += 1
    
    items = sorted(agrup.items(), key=lambda x: (x[0][0], -x[1]['debitos']))
    
    for (tipo, concepto), vals in items:
        ws.cell(row=r, column=1, value=tipo)
        ws.cell(row=r, column=2, value=concepto)
        ws.cell(row=r, column=3, value=vals['cantidad']).alignment = CA
        ws.cell(row=r, column=4, value=vals['debitos']).number_format = MF
        ws.cell(row=r, column=5, value=vals['creditos']).number_format = MF
        _sf(ws, r, 5, DF)
        r += 1

    for c, w in [('A', 22), ('B', 50), ('C', 8), ('D', 18), ('E', 18)]:
        ws.column_dimensions[c].width = w

# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================
def generar_excel(resultado, datos_banco, ruta_salida, mes_anio="", movimientos_sistema=None):
    wb = Workbook()

    _crear_resumen(wb, resultado, datos_banco, mes_anio)
    _crear_auditoria_banco(wb, datos_banco, mes_anio)
    if movimientos_sistema:
        _crear_auditoria_sistema(wb, movimientos_sistema, mes_anio)
        
    _crear_detalle_impuestos(wb, resultado, mes_anio)
    _crear_conciliacion(wb, resultado, datos_banco, mes_anio)
    _crear_conceptos_agrupados(wb, datos_banco, mes_anio)

    wb.save(ruta_salida)
    return ruta_salida
