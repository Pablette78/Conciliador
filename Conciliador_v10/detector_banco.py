"""
detector_banco.py — Auto-detección del banco a partir del archivo de extracto.

Soporta:
  PDF con texto  → Banco Ciudad, Banco Galicia, Banco Comafi, Banco ICBC, Banco Provincia
  PDF imagen     → Banco Santander (sin texto extraíble)
  Excel          → Banco Santander (hoja "descargaUltimosMovimientos")
                 → Banco Galicia u otros Excel (por encabezados)

Retorna uno de:
  "Banco Ciudad" | "Banco Galicia" | "Banco Comafi" | "Banco Santander"
  "Banco ICBC"   | "Banco Provincia" | None (no detectado)

Uso:
    from detector_banco import detectar_banco
    banco = detectar_banco("ruta/al/archivo.pdf")  # → "Banco ICBC"
"""

import os
import re


# ── Fingerprints de texto en PDF ──────────────────────────────────────────────
# Lista ordenada: (banco, [patrones que DEBEN estar presentes])
# Se evalúa de más específico a menos específico.

_FINGERPRINTS_PDF = [
    # ICBC: "IndustrialandCommercialBankofChina" (sin espacios, texto pegado del PDF)
    ("Banco ICBC", [
        r'IndustrialandCommercialBank',
    ]),
    # ICBC fallback: PERIODO DD-MM-YYYY + CUENTA CORRIENTE EN PESOS N°
    ("Banco ICBC", [
        r'PERIODO \d{2}-\d{2}-\d{4}',
        r'CUENTA CORRIENTE EN PESOS N°',
        r'F\.EXT',
    ]),

    # Banco Provincia: "Extracto de Cuenta Informativo" + "Frecuencia MENSUAL"
    # La página de totales tiene "bancoprovincia.com.ar" e "ITF-LEY25413"
    ("Banco Provincia", [
        r'Extracto de Cuenta Informativo',
        r'Frecuencia\s+MENSUAL',
    ]),

    # Banco Ciudad: fechas "DD-ENE-YYYY" + saldo inicial característico
    ("Banco Ciudad", [
        r'BANCO\s+DE\s+LA\s+CIUDAD|Banco\s+Ciudad|BANCO\s+CIUDAD',
    ]),
    ("Banco Ciudad", [
        r'\d{2}-(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)-\d{4}',
        r'SALDO ULTIMO EXTRACTO|SALDO AL \d{2}',
    ]),

    # Banco Galicia
    ("Banco Galicia", [
        r'BANCO GALICIA|Banco Galicia|galicia\.com\.ar',
    ]),
    ("Banco Galicia", [
        r'IMP\. CRE\. LEY 25413|IMP\. DEB\. LEY 25413',
        r'MISMA TITULARIDAD|TRANSFER\. CASH|COELSA',
    ]),

    # Banco Comafi
    ("Banco Comafi", [
        r'COMAFI|Comafi',
    ]),
    # American Express
    ("American Express", [
        r'American Express|AMERICAN EXPRESS',
        r'Estado de Cuenta',
    ]),
    # Tarjeta VISA
    ("Tarjeta VISA", [
        r'VISA BUSINESS|VISA SUPER PYME|Tarjeta VISA|VISA',
        r'TITULAR DE CUENTA|Cuenta:',
    ]),
    # ARCA / AFIP
    ("ARCA-Mis Retenciones", [
        r'AFIP|ARCA',
        r'Mis Retenciones',
    ]),
]


def _texto_pdf_completo(ruta_pdf, max_paginas=6):
    """Extrae y concatena texto de hasta max_paginas del PDF."""
    try:
        import pdfplumber
        textos = []
        tiene_chars = False
        with pdfplumber.open(ruta_pdf) as pdf:
            for pagina in pdf.pages[:max_paginas]:
                texto = pagina.extract_text() or ''
                if texto.strip():
                    textos.append(texto)
                    tiene_chars = True
        return '\n'.join(textos), tiene_chars
    except Exception:
        return '', False


def _es_pdf_imagen(ruta_pdf):
    """Devuelve True si el PDF no tiene texto extraíble (es imagen/OCR)."""
    try:
        import pdfplumber
        with pdfplumber.open(ruta_pdf) as pdf:
            total_chars = sum(len(p.chars) for p in pdf.pages[:3])
            return total_chars == 0
    except Exception:
        return False


def _detectar_desde_excel(ruta_excel):
    """Detecta banco desde un archivo Excel."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta_excel, data_only=True, read_only=True)
        nombre_hoja = wb.sheetnames[0] if wb.sheetnames else ''

        # Santander: hoja "descargaUltimosMovimientos"
        if 'descargaultimos' in nombre_hoja.lower():
            return "Banco Santander"

        # Leer encabezados de la primera hoja
        ws = wb[nombre_hoja]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if cell:
                    headers.append(str(cell).upper())

        texto_headers = ' '.join(headers)

        # Santander Excel: columna "IMPORTE PESOS"
        if 'IMPORTE PESOS' in texto_headers:
            return "Banco Santander"

        # Galicia Excel: columnas típicas de Galicia
        if 'DESCRIPCION' in texto_headers and 'OBSERVACIONES' in texto_headers:
            return "Banco Galicia"

        # Si hay Fecha + Débitos/Créditos → podría ser Galicia u otro
        if 'DEBITOS' in texto_headers or 'DÉBITOS' in texto_headers:
            return "Banco Galicia"

        # AFIP / ARCA
        if 'CUIT AGENTE RET' in texto_headers or 'IMPORTE RET' in texto_headers:
            return "ARCA-Mis Retenciones"

    except Exception:
        pass
    return None


def _detectar_desde_xls(ruta_xls):
    """Detecta banco desde un archivo .xls legacy (usa xlrd)."""
    try:
        import xlrd
        wb = xlrd.open_workbook(ruta_xls)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return None
        # Leer encabezados de las primeras 3 filas
        headers = []
        for r in range(min(3, ws.nrows)):
            for c in range(ws.ncols):
                val = ws.cell_value(r, c)
                if val:
                    headers.append(str(val).upper())
        texto_headers = ' '.join(headers)

        # AFIP / ARCA: encabezados característicos
        if 'CUIT AGENTE RET' in texto_headers or 'IMPORTE RET./PERC.' in texto_headers:
            return "ARCA-Mis Retenciones"

        # Santander .xls
        if 'IMPORTE PESOS' in texto_headers:
            return "Banco Santander"

    except Exception:
        pass
    return None

def detectar_banco(ruta_archivo):
    """
    Detecta automáticamente el banco a partir del archivo de extracto.

    Parámetros:
        ruta_archivo: str — ruta al PDF o Excel del extracto bancario

    Retorna:
        str  — nombre del banco ("Banco Ciudad", "Banco Galicia", etc.)
        None — si no se puede detectar con certeza
    """
    if not os.path.exists(ruta_archivo):
        return None

    extension = os.path.splitext(ruta_archivo)[1].lower()

    # ── Excel ─────────────────────────────────────────────────────────────────
    if extension == '.xls':
        # Los archivos .xls legacy (AFIP/ARCA) requieren xlrd, no openpyxl
        resultado = _detectar_desde_xls(ruta_archivo)
        if resultado:
            return resultado
        # Si xlrd no detectó nada específico, intentar igualmente con openpyxl
        return _detectar_desde_excel(ruta_archivo)

    if extension in ('.xlsx', '.xlsm'):
        return _detectar_desde_excel(ruta_archivo)

    # ── PDF ───────────────────────────────────────────────────────────────────
    if extension != '.pdf':
        return None

    # PDF sin texto → Santander (usa OCR)
    if _es_pdf_imagen(ruta_archivo):
        return "Banco Santander"

    # PDF con texto → comparar fingerprints en todo el documento
    texto, _ = _texto_pdf_completo(ruta_archivo)
    if not texto:
        return None

    for banco, patrones in _FINGERPRINTS_PDF:
        if all(re.search(patron, texto, re.IGNORECASE) for patron in patrones):
            return banco

    return None


def detectar_banco_con_confianza(ruta_archivo):
    """
    Igual que detectar_banco() pero retorna (banco, confianza).
    confianza: "alta" | "media" | "baja" | None
    """
    banco = detectar_banco(ruta_archivo)
    if banco is None:
        return None, None

    extension = os.path.splitext(ruta_archivo)[1].lower()

    # Excel Santander con hoja específica → confianza alta
    if extension in ('.xlsx', '.xls') and banco == "Banco Santander":
        confianza = "alta"
    # PDF imagen → Santander, confianza media (podría ser otro banco escaneado)
    elif extension == '.pdf' and _es_pdf_imagen(ruta_archivo):
        confianza = "media"
    else:
        confianza = "alta"

    return banco, confianza
