
import openpyxl

archivo = r'C:\Pablo Ponti\Conciliador\Conciliador_v10\Ejemplos\Conciliacion_Output_v2.xlsx'
wb = openpyxl.load_workbook(archivo, data_only=True) # data_only=True para ver el resultado de las formulas
ws = wb['Resumen']

print(f"--- Valores en Hoja Resumen ---")
print(f"B16 (Saldo Final Real): {ws['B16'].value}")
print(f"B13 (Saldo Final Ajustado): {ws['B13'].value}")
print(f"B17 (Control Principal): {ws['B17'].value}")
print(f"E8 (Saldo Final Calc D:G): {ws['E8'].value}")
print(f"E10 (Control D:G): {ws['E10'].value}")

ws_conc = wb['4. Conciliación']
print(f"\n--- Controles en Hoja Conciliación ---")
# Buscamos las filas de control al final.
# Según el código, están después de "Solo en Sistema"
for row in range(ws_conc.max_row - 5, ws_conc.max_row + 1):
    lbl = ws_conc.cell(row=row, column=1).value
    dif = ws_conc.cell(row=row, column=7).value
    if lbl and 'Banco' in str(lbl) or 'Sistema' in str(lbl):
        print(f"{lbl}: Diferencia = {dif}")
