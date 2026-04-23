from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ─── Estilos ───────────────────────────────────────────────────────────────────
BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
DF  = Font(name='Arial', size=10)
TF  = Font(name='Arial', bold=True, size=10)
CA  = Alignment(horizontal='center')

AZUL_HEADER = PatternFill('solid', fgColor='2F5496')
AZUL_CLARO  = PatternFill('solid', fgColor='D6E4F0')
VERDE       = PatternFill('solid', fgColor='C6EFCE')
VERDE_TOTAL = PatternFill('solid', fgColor='E2EFDA')
NARANJA     = PatternFill('solid', fgColor='FCE4D6')
AMARILLO    = PatternFill('solid', fgColor='FFEB9C')
ROJO        = PatternFill('solid', fgColor='FFC7CE')
ROJO_FUERTE = PatternFill('solid', fgColor='FF0000')
CELESTE_BOX = PatternFill('solid', fgColor='DDEEFF')
GRIS_CLARO  = PatternFill('solid', fgColor='F2F2F2')

MF  = '#,##0.00'
DTF = 'DD/MM/YYYY'

# Nombres de hojas (constantes para referencias cruzadas)
SH_BANCO = "1. Auditoría Banco"
SH_SIST  = "2. Auditoría Sistema"
SH_IMP   = "3. Detalle Impuestos"
SH_CONC  = "4. Conciliación"


def _h(ws, r, headers, start_col=1):
    for i, h in enumerate(headers):
        col = start_col + i
        cl = ws.cell(row=r, column=col, value=h)
        cl.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cl.fill = AZUL_HEADER
        cl.alignment = CA
        cl.border = BD


def _sf(ws, r, mc, font=None, fill=None):
    for c in range(1, mc + 1):
        cl = ws.cell(row=r, column=c)
        if font: cl.font = font
        if fill: cl.fill = fill
        cl.border = BD


def _titulo(ws, r, texto):
    ws.cell(row=r, column=1, value=texto).font = Font(name='Arial', bold=True, size=12, color='C00000')


def _ref(sheet_name, col, row):
    """Genera fórmula de referencia cruzada a otra hoja."""
    return f"=+'{sheet_name}'!{col}{row}"


def _ctrl_cell(ws, row, col, formula, label_col=None, label=None):
    """Escribe una celda de control con color condicional (verde=0, rojo≠0)."""
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = MF
    c.font = Font(name='Arial', bold=True, size=10)
    c.border = BD
    if label_col and label:
        lc = ws.cell(row=row, column=label_col, value=label)
        lc.font = Font(name='Arial', bold=True, size=10)
        lc.border = BD


# ============================================================
# HOJA 1: AUDITORÍA BANCO
# ============================================================
def _crear_auditoria_banco(wb, datos_banco, mes_anio) -> dict:
    ws = wb.create_sheet(SH_BANCO)

    ws.merge_cells('A1:F1')
    ws['A1'] = f'AUDITORÍA: MOVIMIENTOS DETECTADOS EN BANCO - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    # Saldo anterior (fila 3)
    ws['A3'] = f'Saldo Anterior detectado:'
    ws['A3'].font = Font(name='Arial', bold=True, size=10)
    ws['A3'].border = BD
    ws['D3'] = datos_banco.saldo_anterior or 0
    ws['D3'].number_format = MF
    ws['D3'].font = Font(name='Arial', bold=True, size=10)
    ws['D3'].fill = AMARILLO
    ws['D3'].border = BD
    ROW_SALDO_INI = 3  # columna D

    _h(ws, 5, ['Fecha', 'Concepto Original', 'Débito', 'Crédito', 'Descripción/Ref', 'Clasificación Motor'])
    r = 6
    row_datos_inicio = r
    total_deb = 0.0
    total_cre = 0.0
    for m in datos_banco.movimientos:
        ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=m.concepto)
        if m.debito:
            ws.cell(row=r, column=3, value=m.debito).number_format = MF
            total_deb += m.debito
        if m.credito:
            ws.cell(row=r, column=4, value=m.credito).number_format = MF
            total_cre += m.credito
        ws.cell(row=r, column=5, value=m.descripcion)
        ws.cell(row=r, column=6, value=m.tipo)
        _sf(ws, r, 6, DF)
        r += 1

    row_datos_fin = r - 1

    # Fila TOTALES
    ws.cell(row=r, column=2, value='TOTALES').font = TF
    ws.cell(row=r, column=2).fill = AMARILLO
    ws.cell(row=r, column=2).border = BD
    ws.cell(row=r, column=3, value=f'=SUM(C{row_datos_inicio}:C{row_datos_fin})').number_format = MF
    ws.cell(row=r, column=3).font = TF
    ws.cell(row=r, column=3).fill = AMARILLO
    ws.cell(row=r, column=3).border = BD
    ws.cell(row=r, column=4, value=f'=SUM(D{row_datos_inicio}:D{row_datos_fin})').number_format = MF
    ws.cell(row=r, column=4).font = TF
    ws.cell(row=r, column=4).fill = AMARILLO
    ws.cell(row=r, column=4).border = BD
    for c in [1, 5, 6]:
        ws.cell(row=r, column=c).border = BD
    ROW_TOTALES = r  # C=débitos, D=créditos
    r += 1

    # ── Saldo Final del Extracto (leído por el parser) ──
    # Sirve para verificar que el parser bajó todos los movimientos
    ws.cell(row=r, column=1, value='Saldo Final del Extracto (leído):').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=r, column=1).border = BD
    ws.cell(row=r, column=2, value='Control: Ini + Créd - Déb debe dar este valor →').font = Font(name='Arial', size=9, italic=True, color='666666')
    ws.cell(row=r, column=2).border = BD
    ws.cell(row=r, column=3).border = BD
    ws['D{}'.format(r)] = datos_banco.saldo_final or 0
    ws['D{}'.format(r)].number_format = MF
    ws['D{}'.format(r)].font = Font(name='Arial', bold=True, size=10)
    ws['D{}'.format(r)].fill = AMARILLO
    ws['D{}'.format(r)].border = BD
    ROW_SALDO_FIN = r  # columna D

    for col, w in [('A', 14), ('B', 36), ('C', 16), ('D', 16), ('E', 30), ('F', 24)]:
        ws.column_dimensions[col].width = w

    return {
        'row_saldo_ini': ROW_SALDO_INI,   # D{row} = saldo anterior
        'row_totales':   ROW_TOTALES,      # C{row} = total deb, D{row} = total cred
        'row_saldo_fin': ROW_SALDO_FIN,    # D{row} = saldo final extracto
    }


# ============================================================
# HOJA 2: AUDITORÍA SISTEMA
# ============================================================
def _crear_auditoria_sistema(wb, movimientos_sistema, mes_anio) -> dict:
    ws = wb.create_sheet(SH_SIST)

    ws.merge_cells('A1:F1')
    ws['A1'] = f'AUDITORÍA: MOVIMIENTOS DETECTADOS EN SISTEMA - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    _h(ws, 3, ['Fecha', 'Documento/Ref', 'Concepto Original', 'Detalle', 'Debe', 'Haber'])
    r = 4
    row_datos_inicio = r
    total_debe = 0.0
    total_haber = 0.0
    for m in movimientos_sistema:
        ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
        ws.cell(row=r, column=2, value=m.referencia)
        ws.cell(row=r, column=3, value=m.concepto)
        ws.cell(row=r, column=4, value=m.descripcion)
        if m.debito:
            ws.cell(row=r, column=5, value=m.debito).number_format = MF
            total_debe += m.debito
        if m.credito:
            ws.cell(row=r, column=6, value=m.credito).number_format = MF
            total_haber += m.credito
        _sf(ws, r, 6, DF)
        r += 1

    row_datos_fin = r - 1

    # Fila TOTALES
    ws.cell(row=r, column=4, value='TOTALES').font = TF
    ws.cell(row=r, column=4).fill = AMARILLO
    ws.cell(row=r, column=4).border = BD
    ws.cell(row=r, column=5, value=f'=SUM(E{row_datos_inicio}:E{row_datos_fin})').number_format = MF
    ws.cell(row=r, column=5).font = TF
    ws.cell(row=r, column=5).fill = AMARILLO
    ws.cell(row=r, column=5).border = BD
    ws.cell(row=r, column=6, value=f'=SUM(F{row_datos_inicio}:F{row_datos_fin})').number_format = MF
    ws.cell(row=r, column=6).font = TF
    ws.cell(row=r, column=6).fill = AMARILLO
    ws.cell(row=r, column=6).border = BD
    for c in [1, 2, 3]:
        ws.cell(row=r, column=c).border = BD
    ROW_TOTALES = r  # E=total_debe, F=total_haber

    for col, w in [('A', 14), ('B', 28), ('C', 32), ('D', 32), ('E', 16), ('F', 16)]:
        ws.column_dimensions[col].width = w

    return {'row_totales': ROW_TOTALES}  # E=debe, F=haber


# ============================================================
# HOJA 3: DETALLE DE IMPUESTOS
# ============================================================
def _es_percepcion_iva(concepto: str) -> bool:
    c = concepto.lower()
    return 'percep' in c and ('iva' in c or '2408' in c or 'rg' in c)


def _crear_detalle_impuestos(wb, resultado, mes_anio) -> dict:
    ws = wb.create_sheet(SH_IMP)

    ws.merge_cells('A1:D1')
    ws['A1'] = f'DETALLE DIARIO DE IMPUESTOS Y GASTOS BANCARIOS - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    # Separar percepciones IVA de IVA general
    gastos_expandidos = {}
    for cat, datos in resultado.gastos_por_categoria.items():
        if cat.upper() in ('IVA',):
            items_iva, items_perc = [], []
            for m in datos['items']:
                (items_perc if _es_percepcion_iva(m.concepto) else items_iva).append(m)
            if items_iva:
                gastos_expandidos[cat] = {
                    'total': sum(m.debito or m.credito for m in items_iva),
                    'items': items_iva
                }
            if items_perc:
                gastos_expandidos['IVA Percepción'] = {
                    'total': sum(m.debito or m.credito for m in items_perc),
                    'items': items_perc
                }
        else:
            gastos_expandidos[cat] = datos

    r = 3
    subtotal_rows = []  # para el gran total final
    for cat, datos in sorted(gastos_expandidos.items(), key=lambda x: str(x[0])):
        if len(datos['items']) == 0:
            continue
        nombre_cat = str(cat).replace('_', ' ')
        ws.cell(row=r, column=1, value=nombre_cat).font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws.cell(row=r, column=1).fill = AZUL_HEADER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        _h(ws, r, ['Fecha', 'Concepto Banco', 'Referencia / Detalle', 'Monto'])
        r += 1
        start_r = r
        for m in sorted(datos['items'], key=lambda x: x.fecha):
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            ws.cell(row=r, column=2, value=m.concepto[:80])
            ws.cell(row=r, column=3, value=m.descripcion[:80])
            ws.cell(row=r, column=4, value=m.debito or m.credito).number_format = MF
            _sf(ws, r, 4, DF)
            r += 1
        # Subtotal de categoría
        subtotal_formula = f'=SUM(D{start_r}:D{r - 1})'
        ws.cell(row=r, column=3, value='TOTAL ' + nombre_cat).font = TF
        ws.cell(row=r, column=3).fill = VERDE_TOTAL
        ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=4, value=subtotal_formula).number_format = MF
        ws.cell(row=r, column=4).font = TF
        ws.cell(row=r, column=4).fill = VERDE_TOTAL
        ws.cell(row=r, column=4).border = BD
        for c in [1, 2]:
            ws.cell(row=r, column=c).border = BD
        ws.row_dimensions.group(start_r, r - 1, hidden=True)
        subtotal_rows.append(r)
        r += 2

    # ── GRAN TOTAL de todos los impuestos y gastos ──
    # Es la celda que referencia el bloque A:B del Resumen
    gran_total_formula = '+'.join(f'D{sr}' for sr in subtotal_rows) if subtotal_rows else '0'
    r += 1
    ws.cell(row=r, column=1, value='TOTAL GENERAL IMPUESTOS Y GASTOS').font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws.cell(row=r, column=1).fill = AZUL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=4, value=f'={gran_total_formula}').number_format = MF
    ws.cell(row=r, column=4).font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws.cell(row=r, column=4).fill = AZUL_HEADER
    ws.cell(row=r, column=4).border = BD
    ROW_GRAN_TOTAL = r  # D{row} = total general de impuestos

    for col, w in [('A', 14), ('B', 50), ('C', 45), ('D', 20)]:
        ws.column_dimensions[col].width = w

    return {'row_gran_total': ROW_GRAN_TOTAL}  # D{row} = total impuestos


# ============================================================
# HOJA 4: CONCILIACIÓN
# ============================================================
def _crear_conciliacion(wb, resultado, datos_banco, mes_anio) -> dict:
    ws = wb.create_sheet(SH_CONC)

    ws.merge_cells('A1:M1')
    ws['A1'] = f'CONCILIACIÓN BANCARIA - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = datos_banco.titular
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    # Nota importante sobre la convención de cruce
    ws.merge_cells('A3:M3')
    nota = ws.cell(row=3, column=1,
                   value='⚠  DÉBITOS banco → HABER sistema  ·  CRÉDITOS banco → DEBE sistema  '
                         '— No cruzar las columnas para conciliar')
    nota.font = Font(name='Arial', size=10, italic=True, color='7030A0')
    nota.alignment = CA

    r = 5
    _titulo(ws, r, '1. CRUCE DE MOVIMIENTOS (Banco vs Sistema)')
    r += 1

    _h(ws, r, ['Fecha Banco', 'Déb. Banco', 'Cré. Banco', 'Conc. Banco',
               'Fecha Sist.', 'Debe Sist.', 'Haber Sist.', 'Ref. Sist.',
               'Dif. $', 'Dif. Días', 'Nivel Cierre', 'Estado', 'Alerta/Nota', 'Dif. Cruce'])
    r += 1
    row_conc_inicio = r

    sum_deb_conc_banco  = 0.0
    sum_cre_conc_banco  = 0.0
    sum_debe_conc_sist  = 0.0
    sum_haber_conc_sist = 0.0

    for item in resultado.conciliados:
        b, s = item.banco, item.sistema

        ws.cell(row=r, column=1, value=b.fecha).number_format = DTF
        if b.debito:
            ws.cell(row=r, column=2, value=b.debito).number_format = MF
            sum_deb_conc_banco += b.debito
        if b.credito:
            ws.cell(row=r, column=3, value=b.credito).number_format = MF
            sum_cre_conc_banco += b.credito
        ws.cell(row=r, column=4, value=b.concepto)
        ws.cell(row=r, column=5, value=s.fecha).number_format = DTF
        if s.debito:
            ws.cell(row=r, column=6, value=s.debito).number_format = MF
            sum_debe_conc_sist += s.debito
        if s.credito:
            ws.cell(row=r, column=7, value=s.credito).number_format = MF
            sum_haber_conc_sist += s.credito
        ws.cell(row=r, column=8, value=s.referencia or s.concepto)

        dif_val  = getattr(item, 'diferencia', 0)
        dias_val = getattr(item, 'diferencia_dias', 0)

        dif_cell  = ws.cell(row=r, column=9,  value=dif_val);  dif_cell.number_format = MF
        dias_cell = ws.cell(row=r, column=10, value=dias_val)
        nivel_cell= ws.cell(row=r, column=11, value=item.nivel)
        ws.cell(row=r, column=12, value=item.estado)
        if hasattr(item, 'alerta') and item.alerta:
            ws.cell(row=r, column=13, value=item.alerta).font = Font(name='Arial', color='C00000', bold=True)

        base_fill = VERDE if item.estado == 'CONCILIADO' else ROJO
        _sf(ws, r, 14, DF, base_fill)
        
        # Columna 14: Diferencia Cruce =+B{r}-G{r}-C{r}+F{r}
        dc_cell = ws.cell(row=r, column=14, value=f'=+B{r}-G{r}-C{r}+F{r}')
        dc_cell.number_format = MF; dc_cell.font = DF; dc_cell.border = BD
        if round(dif_val, 2) != 0:
            dif_cell.fill = NARANJA
        if dias_val != 0:
            dias_cell.fill = NARANJA
        if item.nivel != 'FUERTE':
            nivel_cell.fill = NARANJA
        r += 1

    row_conc_fin = r - 1

    # ── SOLO EN BANCO ──
    row_total_sb = None
    if resultado.solo_banco:
        r += 1
        ws.cell(row=r, column=1,
                value=f'SOLO EN BANCO ({len(resultado.solo_banco)} movimientos sin match)').font = Font(
            name='Arial', bold=True, size=11, color='C00000')
        r += 1
        _h(ws, r, ['Fecha', 'Concepto', 'Tipo Automático', 'Débito', 'Crédito', '', '', ''])
        r += 1
        start_sb = r
        for m in resultado.solo_banco:
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            ws.cell(row=r, column=2, value=m.concepto)
            ws.cell(row=r, column=3, value=m.tipo)
            if m.debito:
                ws.cell(row=r, column=4, value=m.debito).number_format = MF
            if m.credito:
                ws.cell(row=r, column=5, value=m.credito).number_format = MF
            _sf(ws, r, 5, DF, NARANJA)
            r += 1

        # Total solo banco
        ws.cell(row=r, column=1, value='TOTAL SOLO EN BANCO').font = TF
        ws.cell(row=r, column=1).fill = AMARILLO
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=4, value=f'=SUM(D{start_sb}:D{r - 1})').number_format = MF
        ws.cell(row=r, column=4).font = TF; ws.cell(row=r, column=4).fill = AMARILLO; ws.cell(row=r, column=4).border = BD
        ws.cell(row=r, column=5, value=f'=SUM(E{start_sb}:E{r - 1})').number_format = MF
        ws.cell(row=r, column=5).font = TF; ws.cell(row=r, column=5).fill = AMARILLO; ws.cell(row=r, column=5).border = BD
        for c in [2, 3]: ws.cell(row=r, column=c).border = BD
        row_total_sb = r  # D=déb, E=cred
        r += 1

    # ── SOLO EN SISTEMA ──
    row_total_ss = None
    if resultado.solo_sistema:
        r += 1
        ws.cell(row=r, column=1,
                value=f'SOLO EN SISTEMA ({len(resultado.solo_sistema)} movimientos sin match)').font = Font(
            name='Arial', bold=True, size=11, color='C00000')
        r += 1
        _h(ws, r, ['Fecha', 'Ref.', 'Concepto', 'Debe', 'Haber', '', '', ''])
        r += 1
        start_ss = r
        for m in resultado.solo_sistema:
            ws.cell(row=r, column=1, value=m.fecha).number_format = DTF
            ws.cell(row=r, column=2, value=m.referencia)
            ws.cell(row=r, column=3, value=m.concepto)
            if m.debito:
                ws.cell(row=r, column=4, value=m.debito).number_format = MF
            if m.credito:
                ws.cell(row=r, column=5, value=m.credito).number_format = MF
            _sf(ws, r, 5, DF, NARANJA)
            r += 1

        # Total solo sistema
        ws.cell(row=r, column=3, value='TOTALES').font = TF
        ws.cell(row=r, column=3).fill = AMARILLO; ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=4, value=f'=SUM(D{start_ss}:D{r - 1})').number_format = MF
        ws.cell(row=r, column=4).font = TF; ws.cell(row=r, column=4).fill = AMARILLO; ws.cell(row=r, column=4).border = BD
        ws.cell(row=r, column=5, value=f'=SUM(E{start_ss}:E{r - 1})').number_format = MF
        ws.cell(row=r, column=5).font = TF; ws.cell(row=r, column=5).fill = AMARILLO; ws.cell(row=r, column=5).border = BD
        for c in [1, 2]: ws.cell(row=r, column=c).border = BD
        row_total_ss = r  # D=debe, E=haber
        r += 1

    # ── CONTROL DE AUDITORÍA ──
    # Verifica: conciliado + no conciliado = total extracto/mayor
    # (si da 0, el parser levantó todo y la conciliación está completa)
    r += 1
    ws.cell(row=r, column=1, value='CONTROL DE AUDITORÍA').font = Font(name='Arial', bold=True, size=11, color='2F5496')
    r += 1
    _h(ws, r, ['Control', 'Descripción', 'Conciliados', 'Solo Banco/Sist.', 'Total', 'Ref. Extracto/Mayor', 'Diferencia (→ 0)'])
    r += 1

    # Para las fórmulas usamos los rangos acumulados
    rng_deb_conc  = f'B{row_conc_inicio}:B{row_conc_fin}'
    rng_cre_conc  = f'C{row_conc_inicio}:C{row_conc_fin}'
    rng_debe_conc = f'F{row_conc_inicio}:F{row_conc_fin}'
    rng_hab_conc  = f'G{row_conc_inicio}:G{row_conc_fin}'

    ref_sb_deb = f'D{row_total_sb}' if row_total_sb else '0'
    ref_sb_cre = f'E{row_total_sb}' if row_total_sb else '0'
    ref_ss_deb = f'D{row_total_ss}' if row_total_ss else '0'
    ref_ss_haber = f'E{row_total_ss}' if row_total_ss else '0'

    controles = [
        ('Banco Déb.',
         'Déb.concil. + Déb.solo banco = Total Déb. Extracto',
         f'=SUM({rng_deb_conc})',
         f'={ref_sb_deb}',
         f'=SUM({rng_deb_conc})+{ref_sb_deb}',
         f"=+'{SH_BANCO}'!C{{rb_tot}}",
         f"=SUM({rng_deb_conc})+{ref_sb_deb}-'{SH_BANCO}'!C{{rb_tot}}"),
        ('Banco Créd.',
         'Cré.concil. + Cré.solo banco = Total Cré. Extracto',
         f'=SUM({rng_cre_conc})',
         f'={ref_sb_cre}',
         f'=SUM({rng_cre_conc})+{ref_sb_cre}',
         f"=+'{SH_BANCO}'!D{{rb_tot}}",
         f"=SUM({rng_cre_conc})+{ref_sb_cre}-'{SH_BANCO}'!D{{rb_tot}}"),
        ('Sistema Debe',
         'Debe concil. + Debe solo sist. = Total Debe Mayor',
         f'=SUM({rng_debe_conc})',
         f'={ref_ss_deb}',
         f'=SUM({rng_debe_conc})+{ref_ss_deb}',
         f"=+'{SH_SIST}'!E{{rs_tot}}",
         f"=SUM({rng_debe_conc})+{ref_ss_deb}-'{SH_SIST}'!E{{rs_tot}}"),
        ('Sistema Haber',
         'Haber concil. + Haber solo sist. = Total Haber Mayor',
         f'=SUM({rng_hab_conc})',
         f'={ref_ss_haber}',
         f'=SUM({rng_hab_conc})+{ref_ss_haber}',
         f"=+'{SH_SIST}'!F{{rs_tot}}",
         f"=SUM({rng_hab_conc})+{ref_ss_haber}-'{SH_SIST}'!F{{rs_tot}}"),
    ]

    # Nota: las referencias a banco/sistema se completarán en generar_excel()
    # Por ahora guardamos los templates; los completamos al llamar a esta función
    # con los row_totales ya conocidos
    _control_rows = []
    for lbl, desc, conc, solo, total, ref, dif in controles:
        ws.cell(row=r, column=1, value=lbl).font = TF
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=desc).font = DF
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=3, value=conc).number_format = MF
        ws.cell(row=r, column=3).font = DF; ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=4, value=solo).number_format = MF
        ws.cell(row=r, column=4).font = DF; ws.cell(row=r, column=4).border = BD
        ws.cell(row=r, column=5, value=total).number_format = MF
        ws.cell(row=r, column=5).font = DF; ws.cell(row=r, column=5).border = BD
        # columna 6 (ref) y 7 (diferencia) se completan luego
        ws.cell(row=r, column=6).border = BD
        dif_c = ws.cell(row=r, column=7)
        dif_c.number_format = MF; dif_c.border = BD
        _control_rows.append(r)
        r += 1

    for col, w in [('A', 14), ('B', 16), ('C', 16), ('D', 28), ('E', 14),
                   ('F', 16), ('G', 16), ('H', 25), ('I', 12), ('J', 10),
                   ('K', 14), ('L', 20), ('M', 35), ('N', 16)]:
        ws.column_dimensions[col].width = w

    return {
        'row_total_sb':    row_total_sb,   # D=déb, E=cred solo banco
        'row_total_ss':    row_total_ss,   # D=debe, E=haber solo sistema
        'control_rows':    _control_rows,  # filas del control de auditoría (cols 6 y 7 pendientes)
        'row_conc_inicio': row_conc_inicio,
        'row_conc_fin':    row_conc_fin,
    }


# ============================================================
# HOJA 5: CONCEPTOS AGRUPADOS
# ============================================================
def _crear_conceptos_agrupados(wb, datos_banco, mes_anio):
    ws = wb.create_sheet("5. Conceptos Agrupados")

    ws.merge_cells('A1:E1')
    ws['A1'] = f'AGRUPACIÓN TOTAL POR CONCEPTOS BANCARIOS - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    agrup = defaultdict(lambda: {'creditos': 0.0, 'debitos': 0.0, 'cantidad': 0})
    for m in datos_banco.movimientos:
        llave = (m.tipo, m.concepto.strip().upper())
        agrup[llave]['creditos'] += m.credito
        agrup[llave]['debitos']  += m.debito
        agrup[llave]['cantidad'] += 1

    r = 3
    _h(ws, r, ['Tipo Clasificación', 'Concepto Exacto', 'Cant.', 'Total Débitos', 'Total Créditos'])
    r += 1
    for (tipo, concepto), vals in sorted(agrup.items(), key=lambda x: (x[0][0], -x[1]['debitos'])):
        ws.cell(row=r, column=1, value=tipo)
        ws.cell(row=r, column=2, value=concepto)
        ws.cell(row=r, column=3, value=vals['cantidad']).alignment = CA
        ws.cell(row=r, column=4, value=vals['debitos']).number_format = MF
        ws.cell(row=r, column=5, value=vals['creditos']).number_format = MF
        _sf(ws, r, 5, DF)
        r += 1

    for col, w in [('A', 22), ('B', 50), ('C', 8), ('D', 18), ('E', 18)]:
        ws.column_dimensions[col].width = w


# ============================================================
# HOJA 0: RESUMEN
# (se llena al final con referencias cruzadas a las otras hojas)
# ============================================================
def _llenar_resumen(ws, resultado, datos_banco, mes_anio,
                    refs_banco, refs_sist, refs_imp, refs_conc):
    """
    Estructura cols A:B — Conciliación bancaria propiamente dicha:

      Saldo Inicial (Banco/Extracto)
    + Total Debe Mayor
    - Total Haber Mayor
    = Saldo Final Calculado (por el Mayor)
    + Créditos no conciliados banco
    - Débitos no conciliados banco
    + Debe no conciliado sistema
    - Haber no conciliado sistema
    - Impuestos y gastos bancarios
    = Saldo Final Ajustado
    ─────────────────────────────
      Saldo Final Real del Extracto
      CONTROL (debe ser 0)

    Estructura cols D:G — "Estado de la descarga del Extracto":
    Verifica que el parser haya capturado todos los movimientos.
      Saldo Inicio  + Créditos - Débitos = Saldo Final Extracto → Control = 0
    """
    ws.title = "Resumen"

    ws.merge_cells('A1:G1')
    ws['A1'] = f'RESUMEN CONCILIACIÓN - {mes_anio}'
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color='2F5496')
    ws['A2'] = datos_banco.titular
    ws['A2'].font = Font(name='Arial', size=10, color='666666')

    # ── Posiciones de referencia ──────────────────────────────────────────────
    rb_ini  = refs_banco.get('row_saldo_ini', 3)      # D{rb_ini} = saldo anterior banco
    rb_tot  = refs_banco.get('row_totales', 7)         # C=deb, D=cred
    rb_sf   = refs_banco.get('row_saldo_fin', 8)       # D{rb_sf} = saldo final extracto
    rs_tot  = refs_sist.get('row_totales', 7)          # E=debe, F=haber  (puede ser None si no hay mayor)
    ri_tot  = refs_imp.get('row_gran_total', 3)        # D{ri_tot} = total impuestos
    r_sb    = refs_conc.get('row_total_sb')            # D=deb, E=cred solo banco
    r_ss    = refs_conc.get('row_total_ss')            # D=debe, E=haber solo sistema

    # Formulas para referencias (si no existe la fila, se usa 0)
    def fref_sist_debe():
        return f"=+'{SH_SIST}'!E{rs_tot}" if refs_sist else '=0'

    def fref_sist_haber():
        return f"=+'{SH_SIST}'!F{rs_tot}" if refs_sist else '=0'

    def fref_sb_deb():
        return f"=+'{SH_CONC}'!D{r_sb}" if r_sb else '=0'

    def fref_sb_cred():
        return f"=+'{SH_CONC}'!E{r_sb}" if r_sb else '=0'

    def fref_ss_debe():
        return f"=+'{SH_CONC}'!D{r_ss}" if r_ss else '=0'

    def fref_ss_haber():
        return f"=+'{SH_CONC}'!E{r_ss}" if r_ss else '=0'

    # ── BLOQUE A:B — Conciliación ─────────────────────────────────────────────
    FONT_NORMAL = Font(name='Arial', size=11)
    FONT_BOLD   = Font(name='Arial', bold=True, size=11)
    FONT_SUBTIT = Font(name='Arial', size=9, italic=True, color='666666')

    def fila_conc(ws, r, etiqueta, formula, negrita=False, fill=None):
        lc = ws.cell(row=r, column=1, value=etiqueta)
        lc.font = FONT_BOLD if negrita else FONT_NORMAL
        lc.border = BD
        if fill: lc.fill = fill
        vc = ws.cell(row=r, column=2, value=formula)
        vc.number_format = MF
        vc.font = FONT_BOLD if negrita else FONT_NORMAL
        vc.border = BD
        if fill: vc.fill = fill

    r = 4
    fila_conc(ws, r, 'Saldo Inicial (Banco / Extracto)',
              f"=+'{SH_BANCO}'!D{rb_ini}", negrita=True, fill=AZUL_CLARO)
    r += 1
    fila_conc(ws, r, '(+) Total Debe Mayor',   fref_sist_debe())
    r += 1
    fila_conc(ws, r, '(-) Total Haber Mayor',  fref_sist_haber())
    r += 1
    # Saldo Final Calculado = B4 + B5 - B6
    fila_conc(ws, r, '(=) Saldo Final Calculado (por el Mayor)',
              '=+B4+B5-B6', negrita=True, fill=AZUL_CLARO)
    r += 1

    # Separador visual
    ws.cell(row=r, column=1, value='── Ajustes por ítems no conciliados ──').font = FONT_SUBTIT
    ws.cell(row=r, column=1).border = BD
    ws.cell(row=r, column=2).border = BD
    r += 1

    fila_conc(ws, r, '(+) Créditos no conciliados banco',  fref_sb_cred())
    r += 1
    fila_conc(ws, r, '(-) Débitos no conciliados banco',   fref_sb_deb())
    r += 1
    fila_conc(ws, r, '(-) Debe no conciliado sistema',     fref_ss_debe())
    r += 1
    fila_conc(ws, r, '(+) Haber no conciliado sistema',    fref_ss_haber())
    r += 1
    fila_conc(ws, r, '(-) Impuestos y gastos bancarios',
              f"=+'{SH_IMP}'!D{ri_tot}")
    r += 1

    # Saldo Final Ajustado = B7 + B8 - B9 + B10 - B11 - B12
    # (posiciones: B7=saldo_calc, B8=cred_nc_banco, B9=deb_nc_banco,
    #              B10=debe_nc_sist, B11=haber_nc_sist, B12=impuestos)
    # Nota: las filas son dinámicas; trabajamos con la posición actual:
    row_saldo_calc   = 7   # fila 7 (B7)
    row_cred_nc_bco  = 9   # fila 9 (B9)  — +1 por el separador visual
    row_deb_nc_bco   = 10
    row_debe_nc_sis  = 11
    row_haber_nc_sis = 12
    row_impuestos    = 13
    # Ajuste de fórmula relativa a B:
    # Signos actualizados: B7 + B8 - B9 - B10 + B11 - B12
    fila_conc(ws, r, '(=) Saldo Final Ajustado',
              f'=+B{row_saldo_calc}+B{row_cred_nc_bco}-B{row_deb_nc_bco}'
              f'-B{row_debe_nc_sis}+B{row_haber_nc_sis}-B{row_impuestos}',
              negrita=True, fill=AZUL_CLARO)
    row_saldo_ajustado = r
    r += 2  # separador

    # Saldo Final Real del Extracto
    ws.cell(row=r, column=1, value='Saldo Final Real del Extracto').font = FONT_BOLD
    ws.cell(row=r, column=1).fill = GRIS_CLARO
    ws.cell(row=r, column=1).border = BD
    ws.cell(row=r, column=2, value=f"=+'{SH_BANCO}'!D{rb_sf}").number_format = MF
    ws.cell(row=r, column=2).font = FONT_BOLD
    ws.cell(row=r, column=2).fill = GRIS_CLARO
    ws.cell(row=r, column=2).border = BD
    row_saldo_final_real = r
    r += 1

    # CONTROL principal → debe ser 0
    ctrl_val = f'=+B{row_saldo_final_real}-B{row_saldo_ajustado}'
    ws.cell(row=r, column=1, value='CONTROL (debe ser 0)').font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=r, column=1).border = BD
    ctrl_cell = ws.cell(row=r, column=2, value=ctrl_val)
    ctrl_cell.number_format = MF
    ctrl_cell.font = Font(name='Arial', bold=True, size=11)
    ctrl_cell.border = BD
    # El color se aplicará condicionalmente (openpyxl no soporta formato condicional fácil,
    # pero sí podemos calcular y aplicar el color en Python)
    v = resultado.validación_saldos
    sfc = v.get('saldo_final_calculado', 0)
    sfe = v.get('saldo_final_extracto', 0) or datos_banco.saldo_final or 0
    ctrl_ok = round(sfe - sfc, 2) == 0
    ctrl_cell.fill = VERDE if ctrl_ok else ROJO_FUERTE
    if not ctrl_ok:
        ctrl_cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        ws.cell(row=r, column=3, value='⚠ Revisar: posibles movimientos sin parsear o sin conciliar').font = Font(
            name='Arial', size=9, color='C00000', italic=True)
    r += 2

    # ── Sección: Estadísticas de conciliación ────────────────────────────────
    ws.cell(row=r, column=1, value='ESTADÍSTICAS DE CONCILIACIÓN').font = Font(name='Arial', bold=True, size=12, color='2F5496')
    r += 1
    total_conc = len(resultado.conciliados)
    con_diff   = sum(1 for c in resultado.conciliados if c.diferencia != 0)
    for desc, val in [
        ('Movimientos conciliados',      total_conc),
        ('Con diferencia de monto',      con_diff),
        ('Solo en banco (pendientes)',   len(resultado.solo_banco)),
        ('Solo en sistema (pendientes)', len(resultado.solo_sistema)),
    ]:
        ws.cell(row=r, column=1, value=desc).font = DF
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=val).font = DF
        ws.cell(row=r, column=2).border = BD
        r += 1

    r += 1
    # Guardamos esta fila para los encabezados compartidos
    row_headers = r
    r += 1
    
    ws.cell(row=r, column=1, value='GASTOS BANCARIOS E IMPUESTOS').font = Font(name='Arial', bold=True, size=12, color='2F5496')
    r += 1
    gran_total_gastos = sum(d['total'] for d in resultado.gastos_por_categoria.values())
    for cat, datos in resultado.gastos_por_categoria.items():
        ws.cell(row=r, column=1, value=cat).font = DF
        ws.cell(row=r, column=1).border = BD
        ws.cell(row=r, column=2, value=datos['total']).number_format = MF
        ws.cell(row=r, column=2).font = DF
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=3, value=f"{len(datos['items'])} movimientos").font = Font(name='Arial', size=9, color='888888')
        r += 1
    ws.cell(row=r, column=1, value='TOTAL GASTOS').font = TF
    ws.cell(row=r, column=1).fill = VERDE_TOTAL
    ws.cell(row=r, column=1).border = BD
    ws.cell(row=r, column=2, value=gran_total_gastos).number_format = MF
    ws.cell(row=r, column=2).font = TF
    ws.cell(row=r, column=2).fill = VERDE_TOTAL
    ws.cell(row=r, column=2).border = BD

    # ── BLOQUE D:G — Estado de la descarga del Extracto ─────────────────────
    # Verifica que el parser capturó todos los movimientos:
    # Saldo Inicio + Créditos - Débitos = Saldo Final Extracto → Control = 0
    fila_box = 4
    col_d, col_e, col_f, col_g = 4, 5, 6, 7

    ws.merge_cells(start_row=fila_box, start_column=col_d, end_row=fila_box, end_column=col_g)
    tit = ws.cell(row=fila_box, column=col_d, value='Estado de la descarga del Extracto')
    tit.font = Font(name='Arial', bold=True, size=11, color='2F5496')
    tit.fill = CELESTE_BOX; tit.border = BD; tit.alignment = CA

    box_items = [
        ('Saldo inicio',  f"=+'{SH_BANCO}'!D{rb_ini}"),
        ('Créditos',      f"=+'{SH_BANCO}'!D{rb_tot}"),
        ('Débitos',       f"=+'{SH_BANCO}'!C{rb_tot}"),
        ('Saldo final (calc.)', '=+E5+E6-E7'),   # E5+E6-E7 relativo
    ]
    for i, (lbl, formula) in enumerate(box_items):
        row_b = fila_box + 1 + i
        cl_lbl = ws.cell(row=row_b, column=col_d, value=lbl)
        cl_lbl.font = DF; cl_lbl.fill = CELESTE_BOX; cl_lbl.border = BD
        ws.merge_cells(start_row=row_b, start_column=col_e, end_row=row_b, end_column=col_f)
        cl_val = ws.cell(row=row_b, column=col_e, value=formula)
        cl_val.number_format = MF; cl_val.font = DF; cl_val.fill = CELESTE_BOX; cl_val.border = BD

    # Control del extracto: saldo calculado vs saldo final real del PDF
    row_ctrl_ext = fila_box + 5
    ctrl_ext_formula = f"=+E{fila_box + 4}-'{SH_BANCO}'!D{rb_sf}"
    ws.cell(row=row_ctrl_ext, column=col_d, value='Control').font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=row_ctrl_ext, column=col_d).fill = CELESTE_BOX
    ws.cell(row=row_ctrl_ext, column=col_d).border = BD
    ws.merge_cells(start_row=row_ctrl_ext, start_column=col_e, end_row=row_ctrl_ext, end_column=col_f)
    ctrl_ext_cell = ws.cell(row=row_ctrl_ext, column=col_e, value=ctrl_ext_formula)
    ctrl_ext_cell.number_format = MF
    ctrl_ext_cell.border = BD

    # Color condicional basado en el valor calculado en Python
    sa  = datos_banco.saldo_anterior or 0
    tc  = sum(m.credito for m in datos_banco.movimientos)
    td  = sum(m.debito  for m in datos_banco.movimientos)
    sfe2 = datos_banco.saldo_final or 0
    ctrl_ext_val = round(sa + tc - td - sfe2, 2)

    if ctrl_ext_val == 0:
        ctrl_ext_cell.font = Font(name='Arial', bold=True, size=11)
        ctrl_ext_cell.fill = VERDE
    else:
        ctrl_ext_cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        ctrl_ext_cell.fill = ROJO_FUERTE

    ws.cell(row=row_ctrl_ext, column=col_g,
            value='* Debe ser 0. Si no, el parser omitió movimientos.').font = Font(
        name='Arial', size=9, color='C00000', italic=True)

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 45

    # ── DUPLICADO: CONTROL DE AUDITORÍA (Mismo que en Conciliación) ──────────
    # Lo posicionamos a la derecha del bloque de estadísticas y gastos
    # Usamos la row_headers que definimos antes para que queden alineados
    ws.cell(row=row_headers - 1, column=col_d, value='AUDITORÍA CRUZADA (Conciliados + Pendientes vs Totales)').font = Font(name='Arial', bold=True, size=11, color='2F5496')
    
    _h(ws, row_headers, ['Control / Categoría', 'Monto', 'Info / Cantidad'], start_col=1)
    _h(ws, row_headers, ['Control Auditoría', 'Referencia Cálculo', 'Suma Celdas', 'Celdas Ref.', 'Diferencia (→ 0)'], start_col=col_d)
    
    r_audit = row_headers
    
    # Definimos los checks de auditoría (ahora usando referencias a las celdas del Resumen A:B y Conciliación)
    auditoria_resumen = [
        ('Banco Déb.', 'Concil + Pend + Imp = Extracto', 
         f"=+B{row_deb_nc_bco}+B{row_impuestos}+SUM('{SH_CONC}'!B{refs_conc['row_conc_inicio']}:B{refs_conc['row_conc_fin']})",
         f"='{SH_BANCO}'!C{rb_tot}"),
        ('Banco Créd.', 'Concil + Pend = Extracto',
         f"=+B{row_cred_nc_bco}+SUM('{SH_CONC}'!C{refs_conc['row_conc_inicio']}:C{refs_conc['row_conc_fin']})",
         f"='{SH_BANCO}'!D{rb_tot}"),
        ('Sist. Debe', 'Concil + Pend = Mayor',
         f"=+B{row_debe_nc_sis}+SUM('{SH_CONC}'!F{refs_conc['row_conc_inicio']}:F{refs_conc['row_conc_fin']})",
         f"='{SH_SIST}'!E{rs_tot}" if rs_tot else '0'),
        ('Sist. Haber', 'Concil + Pend = Mayor',
         f"=+B{row_haber_nc_sis}+SUM('{SH_CONC}'!G{refs_conc['row_conc_inicio']}:G{refs_conc['row_conc_fin']})",
         f"='{SH_SIST}'!F{rs_tot}" if rs_tot else '0'),
    ]

    for lbl, desc, total_calc, ref_valor in auditoria_resumen:
        r_audit += 1
        ws.cell(row=r_audit, column=col_d, value=lbl).font = TF; ws.cell(row=r_audit, column=col_d).border = BD
        ws.cell(row=r_audit, column=col_e, value=desc).font = Font(name='Arial', size=9); ws.cell(row=r_audit, column=col_e).border = BD
        ws.cell(row=r_audit, column=col_f, value=total_calc).number_format = MF; ws.cell(row=r_audit, column=col_f).border = BD
        ws.cell(row=r_audit, column=col_g, value=ref_valor).number_format = MF; ws.cell(row=r_audit, column=col_g).border = BD
        
        # Columna H (8): Diferencia
        col_h = 8
        ws.cell(row=r_audit, column=col_h, value=f'=+F{r_audit}-G{r_audit}').number_format = MF
        ws.cell(row=r_audit, column=col_h).font = Font(name='Arial', bold=True); ws.cell(row=r_audit, column=col_h).border = BD
        ws.cell(row=r_audit, column=col_h).fill = VERDE_TOTAL
    
    ws.column_dimensions[get_column_letter(col_h)].width = 16


# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================
def generar_excel(resultado, datos_banco, ruta_salida, mes_anio="", movimientos_sistema=None):
    wb = Workbook()

    # El primer sheet (activo) lo reservamos para el Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # 1. Generar hojas de datos (retornan posiciones de celdas clave)
    refs_banco = _crear_auditoria_banco(wb, datos_banco, mes_anio)

    refs_sist = {}
    if movimientos_sistema:
        refs_sist = _crear_auditoria_sistema(wb, movimientos_sistema, mes_anio)

    refs_imp  = _crear_detalle_impuestos(wb, resultado, mes_anio)
    refs_conc = _crear_conciliacion(wb, resultado, datos_banco, mes_anio)
    _crear_conceptos_agrupados(wb, datos_banco, mes_anio)

    # 2. Completar las referencias cruzadas en el control de auditoría de Conciliación
    #    (filas 6 y 7 de cada fila de control que quedaron pendientes)
    rb_tot = refs_banco.get('row_totales', 7)
    rs_tot = refs_sist.get('row_totales', 7) if refs_sist else None
    ri_tot = refs_imp.get('row_gran_total', 3)

    refs_ext = [
        (f"=+'{SH_BANCO}'!C{rb_tot}",  f"=SUM(B{refs_conc['row_conc_inicio']}:B{refs_conc['row_conc_fin']})+({'D'+str(refs_conc['row_total_sb']) if refs_conc['row_total_sb'] else '0'})-'{SH_BANCO}'!C{rb_tot}"),
        (f"=+'{SH_BANCO}'!D{rb_tot}",  f"=SUM(C{refs_conc['row_conc_inicio']}:C{refs_conc['row_conc_fin']})+({'E'+str(refs_conc['row_total_sb']) if refs_conc['row_total_sb'] else '0'})-'{SH_BANCO}'!D{rb_tot}"),
        (f"=+'{SH_SIST}'!E{rs_tot}" if rs_tot else '=0', f"=SUM(F{refs_conc['row_conc_inicio']}:F{refs_conc['row_conc_fin']})+({'D'+str(refs_conc['row_total_ss']) if refs_conc['row_total_ss'] else '0'})-'{SH_SIST}'!E{rs_tot}" if rs_tot else '=0'),
        (f"=+'{SH_SIST}'!F{rs_tot}" if rs_tot else '=0', f"=SUM(G{refs_conc['row_conc_inicio']}:G{refs_conc['row_conc_fin']})+({'E'+str(refs_conc['row_total_ss']) if refs_conc['row_total_ss'] else '0'})-'{SH_SIST}'!F{rs_tot}" if rs_tot else '=0'),
    ]

    ws_conc = wb[SH_CONC]
    for i, ctrl_row in enumerate(refs_conc['control_rows']):
        ref_formula, dif_formula = refs_ext[i]
        c6 = ws_conc.cell(row=ctrl_row, column=6, value=ref_formula)
        c6.number_format = MF; c6.font = DF; c6.fill = GRIS_CLARO; c6.border = BD
        c7 = ws_conc.cell(row=ctrl_row, column=7, value=dif_formula)
        c7.number_format = MF; c7.border = BD
        # Color condicional en dif (col 7)
        # No podemos evaluar la fórmula en Python, lo dejamos neutro
        c7.fill = GRIS_CLARO

    # 3. Llenar el Resumen con todas las referencias cruzadas
    _llenar_resumen(ws_resumen, resultado, datos_banco, mes_anio,
                    refs_banco, refs_sist, refs_imp, refs_conc)

    wb.save(ruta_salida)
    return ruta_salida
